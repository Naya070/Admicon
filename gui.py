from operator import and_
from pickle import TRUE
import tkinter as tk
#import tkSimpleDialog
#import tkMessageBox
from tkinter import BOTH, DISABLED, END, RIGHT, Y, StringVar, ttk, Radiobutton
from tkinter import messagebox
from turtle import heading, width
from tkinter import ttk
import tkinter.font as font
from PIL import ImageTk, Image
from base import Controlador
import sqlite3

from datetime import datetime
from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from fpdf import FPDF


LargeFont = ("Verdana", 12)

def options():
     '''\n**************************************************************
          Welcome to the GUI interface of the python program
**************************************************************'''


class PageContainer(tk.Tk):  

	def __init__(self, *args, **kwargs):  
		options()
		tk.Tk.__init__(self, *args, **kwargs) 

		container = tk.Frame(self)  
		tk.Tk.geometry(self,'1440x900')  
		container.pack(side='top', fill='both', expand = True )     
		container.grid_rowconfigure(0, weight = 1)
		container.grid_columnconfigure(0, weight = 1)

		self.frame = {}

		for F in (Login, Control_cuentas, Recibo_factura, Datos_propietarios, Correos_envio, Nomina_empleados):

			frame = F(container, self)

			self.frame[F] = frame

			frame.grid(row = 0, column = 0, sticky = "nsew") 

		self.show_frame(Login)

	def show_frame(self, cont):

		frame = self.frame[cont]    
		frame.tkraise()    



class Login(tk.Frame):    


	def __init__(self, parent, controller):
		tk.Frame.__init__(self, parent)  

		self.fondo = ImageTk.PhotoImage(Image.open("img/fondo.jpg"))
		self.login = tk.PhotoImage(file="img/login.png")
		self.label_fondo = tk.Label(self, image= self.fondo).pack(expand=True)


		label_cuadro = tk.Label(
				self, bg="white", width=50, height=30).place(x=20, y=60)
		label_login = tk.Label(
				self, image= self.login, bg="white").place(x=140, y=60)
		self.ingrese_aqui = tk.Label(self, text="Ingrese Aquí", font=(
				"Arial", 20, "bold"), bg="white", fg="#303452").place(x=110, y=170)
		self.label_usuario = tk.Label(self, text="Usuario o correo", font=(
				"Arial", 15, "bold"), bg="white", fg="#303452").place(x=50, y=230)

		self.entry_usuario_var = StringVar()

		self.entry_usuario = tk.Entry(self, textvariable = self.entry_usuario_var, font=("Arial", 12, "bold")).place(
				x=50, y=270, width=290, height=30)
		self.label_contrase = tk.Label(self, text="Contraseña", font=(
				"Arial", 15, "bold"), bg="white", fg="#303452").place(x=50, y=330)
		
		self.entry_contrase_var = StringVar()
		self.entry_contrase = tk.Entry(self, textvariable = self.entry_contrase_var).place(
				x=50, y=370, width=290, height=30)


		
			#CREAMOS VENTANA PARA REGISTRO.
		def registro():

			ventana_registro = tk.Toplevel()
			ventana_registro.title("Registro")
			ventana_registro.geometry("300x250")
		
			global nombre_usuario
			global clave
			global entrada_nombre
			global entrada_clave
			nombre_usuario = StringVar() #DECLARAMOS "string" COMO TIPO DE DATO PARA "nombre_usuario"
			clave = StringVar() #DECLARAMOS "sytring" COMO TIPO DE DATO PARA "clave"
		
			tk.Label(ventana_registro, text="Introduzca datos", bg="LightGreen", font=("Arial", 12, "bold")).pack()
			tk.Label(ventana_registro, text="").pack()
			etiqueta_nombre = tk.Label(ventana_registro, text="Nombre de usuario * ",font=("Arial", 12, "bold"))
			etiqueta_nombre.pack()
			entrada_nombre = tk.Entry(ventana_registro, textvariable=nombre_usuario, font=("Arial", 12, "bold")) #ESPACIO PARA INTRODUCIR EL NOMBRE.
			entrada_nombre.pack()
			etiqueta_clave = tk.Label(ventana_registro, text="Contraseña * ",font=("Arial", 12, "bold"))
			etiqueta_clave.pack()
			entrada_clave = tk.Entry(ventana_registro, textvariable=clave, show='*',font=("Arial", 12, "bold")) #ESPACIO PARA INTRODUCIR LA CONTRASEÑA.
			entrada_clave.pack()
			tk.Label(ventana_registro, text="").pack()
			tk.Button(ventana_registro, text="Registrarse", width=10, height=1, bg="LightGreen", command = registro_usuario).pack() #BOTÓN "Registrarse"
		
		def comprobar_datos():

			if len(self.entry_usuario_var.get()) < 1:
				messagebox.showinfo("ADVERTENCIA","Coloque el nombre de usuario")
			
			elif len(self.entry_contrase_var.get()) < 1:
				messagebox.showinfo("ADVERTENCIA","Coloque la contraseña")
			
			else:
				self.my_connexion = sqlite3.connect("Admicon.db")
				cursor = self.my_connexion.cursor()
				self.datos = cursor.execute("SELECT * FROM Login WHERE user = '%s'" % (self.entry_usuario_var.get()))
				for row in self.datos:
					self.elusuario = row[1]
					self.lacontra =  row[2]
				
				if self.elusuario != self.entry_usuario_var.get():
					messagebox.showinfo("ADVERTENCIA","Usuario incorrecto")

				elif self.lacontra != self.entry_contrase_var.get():
					messagebox.showinfo("ADVERTENCIA","Contraseña incorrecta")

				else:
					controller.show_frame(Control_cuentas)

				
				

		self.button_acceder = tk.Button(self, text="Acceder", command = comprobar_datos, font=(
				"Arial", 15, "bold"), bg="#303452", fg="white", width=15, height=2).place(x=100, y=430)
		self.button_registrar = tk.Button(self, text="Registrarse", command= registro, font=(
				"Arial", 15, "bold"), bg="white", fg="#303452", width=15, height=2).place(x=650, y=430)

		#REGISTRO DE USUARIO
		def registro_usuario():
			self.my_connexion = sqlite3.connect("Admicon.db")
			cursor = self.my_connexion.cursor()
		
			try:

				self.usuario = nombre_usuario.get()
				print(self.usuario)
				self.contrasena = clave.get()
				print(self.contrasena)

				cursor.execute("INSERT INTO Login VALUES (NULL,'%s','%s')" % (self.usuario, self.contrasena))
				self.my_connexion.commit()
				messagebox.showinfo("ADVERTENCIA","Se guardó el recibo con éxito")

			except:
				messagebox.showwarning("ADVERTENCIA","Ocurrió un error al registrar usuario.")
				pass

		
		
		


class Control_cuentas(tk.Frame):

	controlbd = Controlador()

	def __init__(self, parent, controller):

		tk.Frame.__init__(self, parent)
			

		self.frame_fondo = tk.Frame(self)
		self.frame_fondo.pack(expand=True)
		self.frame_fondo.config(bg="white", width=1440, height=300)

		self.frame_1 = tk.Frame(self.frame_fondo)
		self.frame_1.pack(fill=tk.BOTH, side=tk.TOP)
		self.frame_1.config(bg="light blue", width=1440, height=300)

		self.label_a= tk.Label(self.frame_1,bg="light blue", relief=tk.SUNKEN)
		self.label_a.place(x=10,y=40, width=280, height= 250)

		self.label_b= tk.Label(self.frame_1,bg="light blue", relief=tk.SUNKEN)
		self.label_b.place(x=295,y=90, width=360, height= 170)

		self.label_c= tk.Label(self.frame_1,bg="light blue", relief=tk.SUNKEN)
		self.label_c.place(x=660,y=90, width=360, height= 170)

		self.button_recibo = tk.Button(self, text="Recibos", 
		command = lambda:controller.show_frame(Recibo_factura), width=20, height=1).place(x=0, y=0)

		self.button_propietarios = tk.Button(self, text="Datos Propietarios", 
		command = lambda:controller.show_frame(Datos_propietarios), width=20, height=1).place(x=420, y=0)
		self.button_correos = tk.Button(self, text="Envío Correos",
		command = lambda:controller.show_frame(Correos_envio), width=20, height=1).place(x=630, y=0)
		self.button_nomina = tk.Button(self, text="Nómina", 
		command = lambda:controller.show_frame(Nomina_empleados), state= DISABLED, width=20,
								height=1).place(x=840, y=0)

		#registrar_pago =  Button(self, text="Registrar Pago", command= registro_pago, width=20, height=1).place(x = 630, y = 60)
		self.añadir_recibo =  tk.Button(self, text="Añadir monto de recibo", command= self.añadir_monto_recibo, width=20, height=1).place(x = 70, y = 160)

		self.bolivares_acumulados = tk.Label(self, text="Total Bs:", font=("Arial"), bg="lightblue").place(x=18, y=210)
		self.bolivares_acumulados_mostrar= tk.Label(self.frame_1,  bg="light blue", relief=tk.SUNKEN, font=("Arial")).place(x=120,y=203, width=150)
		
		self.dolares_acumulados = tk.Label(self, text="Total $:", font=("Arial"), bg="lightblue").place(x=18, y=250)
		self.dolares_acumulados_mostrar= tk.Label(self.frame_1, bg="light blue", relief=tk.SUNKEN, font=("Arial")).place(x=120,y=250, width=150)

		self.buscar = tk.Button(self, text="Buscar", command = self.busqueda, width=10, height=1).place(x = 860, y = 38)

		#instancia = Recibo_factura()
		#resultado = instancia.get_monto_recibo()
		#print(Recibo_factura.get_monto_recibo)
		cuentas_consulta_apto = self.controlbd.consulta_total()
		self.m_recibo = cuentas_consulta_apto[0][1]
		self.m_recibo_coma = str(self.m_recibo).replace('.',',')
		

		self.label_button_cuentas = tk.Label(self, text="Control de Cuentas", font=(
				"Arial"), bg="lightblue", width=20, height=1).place(x=210, y=0)
		self.total_pago = tk.Label(self, text="Total Pago:", font=(
				"Arial"), bg="lightblue", width=10, height=1).place(x=20, y=60)
		self.total_pago_mostrar = tk.Label(self, text= str(self.m_recibo_coma), font=(
				"Arial"), bg="lightblue", relief=tk.SUNKEN).place(x=120, y=60, width=150)

		self.porcentaje_mora = tk.Label(self, text="1 % Mora:", font=(
				"Arial"), bg="lightblue", width=10, height=1).place(x=20, y=90)
		self.porcentaje_mora_mostrar = tk.Label(self, text="12345678,90 Bs", font=(
				"Arial"), bg="lightblue", relief=tk.SUNKEN).place(x=120, y=90, width=150)
		self.fecha_self = tk.Label(self, text="Fecha:", font=("Arial"),
							bg="lightblue", width=10, height=1).place(x=20, y=120)
		self.fecha_self_mostrar = tk.Label(self, text= datetime.now().strftime('%d-%m-%Y'),font=("Arial"),
									bg="lightblue", relief=tk.SUNKEN).place(x=120, y=120, width=120)

		self.registrar_pago = tk.Label(self, text="Registrar pago", font=(
				"Arial",), bg="lightblue", fg="gray44", width=13, height=1).place(x=310, y=80)
		self.añadir_cobro_etiqueta = tk.Label(self, text="Cobro alquiler de estacionamiento", font=(
				"Arial",), bg="lightblue", fg="gray44", width=30, height=1).place(x=680, y=80)
		
		self.buscar = tk.Label(self, text="Buscar:", font=("Arial"),
							bg="lightblue", width=8, height=1).place(x=600, y=43)

		self.buscar_entry_var = StringVar()
		self.buscar_entry = tk.Entry(self, textvariable= self.buscar_entry_var).place(x=680, y=44, width=170)

		self.apartamento_pago= tk.Label(self, text="Apartamento:", font=("Arial"), bg="lightblue" ).place(x=300, y=43)  
		self.acumulado= tk.Label(self, text="Acumulado:", font=("Arial"), bg="lightblue" ).place(x=300, y=110) 
		
		self.monto_pago= tk.Label(self, text="Bolívares:", font=("Arial"), bg="lightblue" ).place(x=300, y=150) 
		self.monto_pago_usd= tk.Label(self, text="Dólares:", font=("Arial"), bg="lightblue" ).place(x=490, y=150) 

		self.monto_cobro= tk.Label(self, text="Bolívares:", font=("Arial"), bg="lightblue" ).place(x=685, y=150) 
		#self.monto_cobro_usd= tk.Label(self, text="Dólares:", font=("Arial"), bg="lightblue" ).place(x=855, y=150) 
		self.restante= tk.Label(self, text="Restante:", font=("Arial"), bg="lightblue" ).place(x=500, y=275)
		#self.saldo= tk.Label(self, text="Saldo:", font=("Arial"), bg="lightblue" ).place(x=665, y=265) 

		#self.apartamento_cobro= tk.Label(self, text="Apartamento:", font=("Arial"), bg="lightblue" ).place(x=665, y=105) 


		self.apartamento_entmonto = StringVar()
		self.apartamento_entry_pago = tk.Entry(self, textvariable= self.apartamento_entmonto).place(x=410, y=44, width=170)
		self.apartamento_entry_pago = tk.Label(self, text= "",font=("Arial"), bg="lightblue", relief=tk.SUNKEN).place(x=410, y=44, width=170)
		
		self.monto_entry_pago_bs_var = StringVar()
		self.monto_entry_pago_bs = tk.Entry(self, textvariable=self.monto_entry_pago_bs_var).place(x=300, y=173, width=150)

		self.monto_entry_pago_usd_var = StringVar()
		self.monto_entry_pago_usd = tk.Entry(self, textvariable=self.monto_entry_pago_usd_var).place(x=490, y=173, width=150)


		
		self.monto_entry_cobro_var = StringVar()
		self.monto_entry_cobro = tk.Entry(self, textvariable=self.monto_entry_cobro_var).place(x= 785, y=153, width=150)

		#self.monto_entry_cobro_usd_var = StringVar()
		#self.monto_entry_cobro_usd = tk.Entry(self, textvariable=self.monto_entry_cobro_var).place(x=855, y=173, width=150)
		
		#self.apartamento_entry_cobro_var = StringVar()
		#self.apartamento_entry_cobro = tk.Entry(self, textvariable=self.apartamento_entry_cobro_var).place(x=805, y=108, width=200)

		self.etiqueta_pago_acumulado = "0,0"
		self.etiqueta_pago_monto = "0,0"
		self.etiqueta_pago_restante = "0,0"

		#self.etiqueta_cobro_monto = "0,0"
		#self.etiqueta_cobro_saldo = "0,0"

		self.acumulado_label = tk.Label(self, text = self.etiqueta_pago_acumulado, font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=441, y=113, width=198) 
		self.Monto_Bs_pago = tk.Label(self, text= self.etiqueta_pago_monto, font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=491, y=193, width=149) 
		self.restante_mostrar_pago = tk.Label(self, text=self.etiqueta_pago_restante, font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=600, y=275, width=200) 

		#self.a = tk.IntVar()
		#self.a.set(value=1)
		#Radiobutton(self, text="Bs", variable=self.a, value=1, bg="lightblue" ).place(x=355, y=175)
		#Radiobutton(self, text="$", variable=self.a, value=2, bg="lightblue" ).place(x=395, y=175)

		
		#self.acumulado_mostrar =  tk.Button(self, text="Mostrar Acumulado",  width=15, height=1).place(x = 350, y = 60)
		self.añadir_pago_button =  tk.Button(self, text="Añadir", command= self.añadir_pago , width=10, height=1).place(x = 420, y = 230)

		self.cba = tk.StringVar() #variable var entero
		self.c = tk.Checkbutton(self, text= "Añadir a Todos los Apartamentos", variable= self.cba, onvalue ="1", offvalue="2", font=("Arial"), bg="lightblue")
		self.c.deselect()
		self.c.place(x=665, y=110)
		

		#self.Monto_Bs_cobro = tk.Label(self, text= self.etiqueta_cobro_monto, font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=856, y=193, width=149) 
		#self.restante_mostrar = tk.Label(self, text = self.etiqueta_cobro_saldo, font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=805, y=265, width=200) 

		#self.a2 = tk.IntVar()
		#self.a2.set(value=1)
		#Radiobutton(self, text="Bs", variable=self.a2, value=1, bg="lightblue" ).place(x=720, y=175)
		#Radiobutton(self, text="$", variable=self.a2, value=2, bg="lightblue" ).place(x=765, y=175)

		self.añadir_cobro_button =  tk.Button(self, text="Añadir", command= self.añadir_cobro, width=10, height=1).place(x = 790, y = 230)


			
		# Frame del treeview
		self.frame_treeview = tk.Frame(self.frame_fondo)
		self.frame_treeview.pack(fill=tk.BOTH, side=tk.BOTTOM)
		self.frame_treeview.config(bg="white", width=2440, height=300)


				# Set the treeview
		self.tree = ttk.Treeview(self.frame_treeview)

		self.treexscroll = tk.Scrollbar(self.frame_treeview, orient=tk.HORIZONTAL)
		self.treexscroll.pack(fill=tk.X, side=tk.BOTTOM)
		# configurar scrollbar
		self.treexscroll.config(command=self.tree.xview)

		self.treeyscroll = tk.Scrollbar(self.frame_treeview, orient=tk.VERTICAL)
		self.treeyscroll.pack(fill=tk.Y, side=tk.RIGHT)
		# configurar scrollbar
		self.treeyscroll.config(command=self.tree.yview)

		# TREEVIEW
		self.tree.config(xscrollcommand=self.treexscroll.set, yscrollcommand=self.treeyscroll.set,
		columns=(
				"col1", "col2", "col3", "col4", "col5", "col6", "col7", "col8", "col9", "col10", "col11", "col12"))

	
		self.tree.column("#0", width= 50, stretch= False)
		self.tree.column("col1", width=150, stretch= False)
		self.tree.column("col2", width=150, stretch= False)
		self.tree.column("col3", width=150, stretch= False)
		self.tree.column("col4", width=160, stretch= False)
		self.tree.column("col5", width=150, stretch= False)
		self.tree.column("col6", width=150, stretch= False)
		self.tree.column("col7", width=150, stretch= False)
		self.tree.column("col8", width=150, stretch= False)
		self.tree.column("col9", width=150, stretch= False)
		self.tree.column("col10", width=150, stretch= False)
		self.tree.column("col11", width=150, stretch= False)
		self.tree.column("col12", width=150, stretch= False)

		self.tree.heading("#0", text="ID", anchor=tk.CENTER)
		self.tree.heading("col1", text="Apartamento", anchor=tk.CENTER)
		self.tree.heading("col2", text="Deudas del mes pasado", anchor=tk.CENTER)
		self.tree.heading("col3", text="Recibo monto", anchor=tk.CENTER)
		self.tree.heading("col4", text="Mora", anchor=tk.CENTER)
		self.tree.heading("col5", text="Alquiler estacionamiento", anchor=tk.CENTER)
		self.tree.heading("col6", text="Deuda actual", anchor=tk.CENTER)
		self.tree.heading("col7", text="Bolívares", anchor=tk.CENTER)
		self.tree.heading("col8", text="Dólares", anchor=tk.CENTER)
		self.tree.heading("col9", text="Fecha", anchor=tk.CENTER)
		self.tree.heading("col10", text="Cambio del día", anchor=tk.CENTER)
		self.tree.heading("col11", text="Dólares a Bs", anchor=tk.CENTER)
		self.tree.heading("col12", text="Saldo", anchor=tk.CENTER)


		self.tree.pack()

	
		self.id = 0
		self.iid = 0

		cuentas_apt = self.controlbd.consulta_cuentas()
		
		#MODIFICAR PARA QUE IMPRIMA UNA SOLA VEZ
		self.indice = 0
		while self.indice < len(cuentas_apt):
			self.tree.tag_configure("gray90", background="gray90")
			self.tree.tag_configure("white", background="white")
			color = "white" if self.indice % 2 else "gray90"
			self.a = cuentas_apt[self.indice][6]
			self.b = cuentas_apt[self.indice][12]

			if cuentas_apt[self.indice][6] != None:
				self.a = round(cuentas_apt[self.indice][6], 2)
			
			if cuentas_apt[self.indice][12] != None:
				self.b = round(cuentas_apt[self.indice][12],2)
			
			self.tree.insert("",END, tag=('fuente', color),  iid=self.indice, text= cuentas_apt[self.indice][0], 
			values =( cuentas_apt[self.indice][1],(str(cuentas_apt[self.indice][2])).replace('.',','), 
					(str(cuentas_apt[self.indice][3])).replace('.',','),(str(cuentas_apt[self.indice][4])).replace('.',','),
					(str(cuentas_apt[self.indice][5])).replace('.',','),(str(self.a)).replace('.',','),
					(str(cuentas_apt[self.indice][7])).replace('.',','),(str(cuentas_apt[self.indice][8])).replace('.',','),
					cuentas_apt[self.indice][9],(str(cuentas_apt[self.indice][10])).replace('.',','), 
					(str(cuentas_apt[self.indice][11])).replace('.',','), (str(self.b)).replace('.',',')))

			self.indice = self.indice +1

		

		#((str(round(r[2], 2))).replace('.',',')
		self.tree.bind('<<TreeviewSelect>>', self.seleccionarUsandoClick)
		self.tree.bind('<<TreeviewSelect>>', self.bindings)


		#SUMA Y RESTA DE COLUMNAS EN BASE DE DATOS
		#UPDATE tabla SET  colum3 = (colum1 - colum2);
		self.my_connexion = sqlite3.connect("Admicon.db")
		cursor = self.my_connexion.cursor()
		cursor.execute("UPDATE Cuentas_por_apartamento SET  Deuda_actual = ( Deudas_mes_pasado + Recibo + Mora + Alquiler_estacionamiento)")
		self.my_connexion.commit()
		cursor.execute("UPDATE Cuentas_por_apartamento SET  Saldo = (Deuda_actual - Pago_Bs - Dolar_a_Bs)")
		self.my_connexion.commit()

		self.total_bs_usd()


	def seleccionarUsandoClick(self, event):
			item = self.tree.identify('item', event.x, event.y)
			self.apartamento_entmonto.set(self.tree.item(item, "values")[0])
			#if self.a.get()==1:
			self.monto_entry_pago_bs_var.set(self.tree.item(item, "values")[6])
			#elif self.a.get()==2:
			self.monto_entry_pago_usd_var.set(self.tree.item(item, "values")[7])
			#self.apartamento_entry_cobro_var.set(self.tree.item(item, "values")[0])
			self.monto_entry_cobro_var.set(self.tree.item(item, "values")[4])

			print("you clicked on", self.tree.item(item,"text"))
			self.id_c = self.tree.item(item,"text")
			#self.tree.selection_set('0')

			
			self.etiqueta_pago_acumulado = round(float((self.tree.item(item,"values")[5]).replace(',','.')), 2)
			self.etiqueta_pago_acumulado = (str(self.etiqueta_pago_acumulado)).replace('.',',')
			
			self.etiqueta_pago_monto = round(float((self.tree.item(item,"values")[10]).replace(',','.')), 2)
			self.etiqueta_pago_monto = (str(self.etiqueta_pago_monto)).replace('.',',')
			
			self.etiqueta_pago_restante = round(float((self.tree.item(item,"values")[11]).replace(',','.')), 2)
			self.etiqueta_pago_restante = (str(self.etiqueta_pago_restante)).replace('.',',')
			
			
			self.etiqueta_cobro_saldo = round(float((self.tree.item(item,"values")[11]).replace(',','.')), 2)
			self.etiqueta_cobro_saldo = (str(self.etiqueta_cobro_saldo)).replace('.',',')

			self.apartamento_entry_pago = tk.Label(self, text= self.tree.item(item, "values")[0],font=("Arial"), bg="lightblue", relief=tk.SUNKEN).place(x=410, y=44, width=170)
			
			self.acumulado_label = tk.Label(self, text = self.etiqueta_pago_acumulado, font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=441, y=113, width=198) 
			self.Monto_Bs_pago = tk.Label(self, text = self.etiqueta_pago_monto, font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=491, y=193, width=149) 
			self.restante_mostrar_pago = tk.Label(self, text=self.etiqueta_pago_restante, font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=600, y=275, width=200) 

			#self.Monto_Bs_cobro = tk.Label(self, text= self.etiqueta_cobro_monto, font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=806, y=193, width=198) 
			#self.restante_mostrar = tk.Label(self, text = self.etiqueta_cobro_saldo, font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=805, y=265, width=200) 

	def total_bs_usd(self):
		self.my_connexion = sqlite3.connect("Admicon.db")
		cursor = self.my_connexion.cursor()

		self.datos = cursor.execute("SELECT Pago_Bs, Pago_USD FROM Cuentas_por_apartamento")
		self.bs= 0
		self.usd= 0
		for row in self.datos:
			print(row)
			if row[0] != None:
				self.bs= self.bs+row[0]
				print(self.bs)
			if row[1] != None:
				self.usd= self.usd+row[1]
				print(self.usd)
			
		
		cursor.execute("UPDATE Total_apto SET Total_bs='%s', Total_usd='%s' WHERE id=1"% (self.bs, self.usd))
		self.my_connexion.commit()
		
		self.bolivares_acumulados_mostrar= tk.Label(self.frame_1, text= str(self.bs).replace('.',','), bg="light blue", relief=tk.SUNKEN, font=("Arial")).place(x=120,y=203, width=150)
		self.dolares_acumulados_mostrar= tk.Label(self.frame_1, text= str(self.usd).replace('.',','), bg="light blue", relief=tk.SUNKEN, font=("Arial")).place(x=120,y=250, width=150)
	
	def bindings(self, event):
			self.tree.bind("<Button-1>", self.seleccionarUsandoClick)

	def cambiar_etiqueta_total_pago(self):
		self.total_pago_mostrar = tk.Label(self, text= self.m_recibo_coma, font=("Arial"), bg="lightblue", relief=tk.SUNKEN).place(x=120, y=60, width=150)

	def añadir_pago(self):
		self.my_connexion = sqlite3.connect("Admicon.db")
		cursor = self.my_connexion.cursor()
		try:
			print("A")
			self.datos = cursor.execute("SELECT Dolar_escogido FROM Total_apto WHERE id=1")
			for row in self.datos:
				self.row_id = row[0]
			print("B")
			self.pago_usd = float(self.monto_entry_pago_usd_var.get().replace(',','.'))
			self.dolares_a_bolivares = self.row_id*self.pago_usd
			self.dolares_a_bolivares = round(self.dolares_a_bolivares, 2)
		
			print(self.pago_usd)
			print(self.dolares_a_bolivares)
			print("C")
			#.replace('.',',')
		
			self.criterio1= round(float(str(self.monto_entry_pago_bs_var.get()).replace(',','.')), 2)
			self.criterio2= round(float(str(self.monto_entry_pago_usd_var.get()).replace(',','.')), 2)
			self.criterio3= datetime.now().strftime('%d-%m-%Y')
			self.criterio4= self.row_id
			self.criterio5= self.dolares_a_bolivares
			self.criterio6= self.apartamento_entmonto.get()
			print("D")
			print(self.criterio1, self.criterio2, self.criterio3, self.criterio4, self.criterio5, self.criterio6)

		

			cursor.execute("UPDATE Cuentas_por_apartamento SET Pago_Bs='%s', Pago_USD='%s', Fecha='%s', Cambio='%s', Dolar_a_Bs='%s' WHERE Apartamento='%s'"% (self.criterio1, self.criterio2, self.criterio3, self.criterio4, self.criterio5, self.criterio6))
			self.my_connexion.commit()
			print("E")
			
			self.total_bs_usd()
			self.sumar_restar_columnas()
			self.mostrar()
			self.limpiar_campos_pago()
			

		except:
			messagebox.showwarning("ADVERTENCIA","Ocurrió un error al añadir pago")
			pass
		

	
	def añadir_cobro_codigo(self):
			self.my_connexion = sqlite3.connect("Admicon.db")
			print(111)
			self.cursor = self.my_connexion.cursor()
			print(2)
			self.monto_cobro_float = float(self.monto_entry_cobro_var.get().replace(',','.'))
			print(3)
			self.monto_cobro_float = round(self.monto_cobro_float, 2)
			print(4)
		
			self.checkbutton_c = self.cba.get()
			print(5)
		
			if self.checkbutton_c == "1": #seleccionado
				answer = messagebox.askokcancel(message="Se añadirá cobro del alquiler de estacionamiento a todos los apartamentos ¿Desea continuar?", title="Confirmación")

				if answer is True:
					print(6)
					self.cursor.execute("UPDATE Cuentas_por_apartamento SET Alquiler_estacionamiento='%s'"% (self.monto_cobro_float))
					print(7)
					self.my_connexion.commit()
				elif answer is False:
					pass
				else:
					print("error")
					print(8)
				
				

			elif self.checkbutton_c == "2": #No seleccionado
				print(9)
				self.cursor.execute("UPDATE Cuentas_por_apartamento SET Alquiler_estacionamiento='%s' WHERE Apartamento='%s'"% (self.monto_cobro_float, self.apartamento_entmonto.get()))	
				print(10)
				self.my_connexion.commit()
				print(11)
				self.datos = self.cursor.execute("SELECT Saldo FROM Cuentas_por_apartamento WHERE Apartamento='%s'"% (self.apartamento_entmonto.get()))
				print(12)
				for row in self.datos:
					print(13)
					self.row_id = str(row[0]).replace('.',',')
					print(14)
				
				self.datos = self.cursor.execute("SELECT Saldo FROM Cuentas_por_apartamento WHERE Apartamento='%s'"% (self.apartamento_entmonto.get()))
				print(15)
				for row in self.datos:
					print(16)
					self.row_id = str((round(row[0], 2))).replace('.',',')
					print(17)
					
				
				#self.restante_mostrar = tk.Label(self, text = self.row_id, font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=805, y=265, width=200)
				print(18) 
			
			self.sumar_restar_columnas()
			print(19)
			self.mostrar()
			print(20)
			self.limpiar_campos_cobro()
			print(21)	
		

	def añadir_cobro(self):
		try:
			self.añadir_cobro_codigo()
		except:
			messagebox.showwarning("ADVERTENCIA","Ocurrió un error al añadir cobro")
			pass


	def limpiar_campos_pago(self):
		try:
			self.apartamento_entmonto.set("")
			self.monto_entry_pago_bs_var.set("")
			self.monto_entry_pago_usd_var.set("")

			self.acumulado_label = tk.Label(self, text = "0,0", font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=441, y=113, width=198) 
			self.Monto_Bs_pago = tk.Label(self, text= "0,0", font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=491, y=193, width=149) 
			self.restante_mostrar_pago = tk.Label(self, text= "0,0", font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=600, y=275, width=200)
			#self.restante_mostrar = tk.Label(self, text = "0,0", font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=600, y=275, width=200)

		except:
			messagebox.showwarning("ADVERTENCIA","Ocurrió un error al limpiar campos de pago")
			pass

	def limpiar_campos_cobro(self):
		try:
			self.monto_entry_cobro_var.set("")

			self.restante_mostrar_pago = tk.Label(self, text= "0,0", font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=600, y=275, width=200)
			self.restante_mostrar = tk.Label(self, text = "0,0", font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=600, y=275, width=200)

		except:
			messagebox.showwarning("ADVERTENCIA","Ocurrió un error al limpiar campos de pago")
			pass
	
	def busqueda(self):
			self.my_connexion = sqlite3.connect("Admicon.db")
			cursor = self.my_connexion.cursor()

		
			try:
				self.criterio = self.buscar_entry_var.get()
				self.criterio = "%s" % self.criterio +"%"
				self.datos = cursor.execute("SELECT id FROM Cuentas_por_apartamento WHERE Apartamento LIKE '%s'" % (self.criterio))
				for row in self.datos:
					self.row_id = row[0]-1
				print(self.row_id)
				self.tree.selection_set(self.tree.tag_has(self.row_id)) # move selection
				self.tree.selection_set(self.row_id) # move selection
				self.tree.focus(self.row_id) # move focus
				self.tree.see(self.row_id) # scroll to show it
        	
			
			except:
				messagebox.showwarning("ADVERTENCIA","Ocurrió un error de búsqueda")
				pass


	def mostrar(self): #Actualizar treeview luego de modificar
		registros = self.tree.get_children()
		print("a")
		for elemento in registros:
			self.tree.delete(elemento)
			
		try:
			print("d")
			
			cuentas_apt = self.controlbd.consulta_cuentas()
			print("e")
			self.indice = 0
			print("f")
			print("indice: ", self.indice)
			print("len: ", len(cuentas_apt))
			while self.indice < len(cuentas_apt):
				self.tree.tag_configure("gray90", background="gray90")
				self.tree.tag_configure("white", background="white")
				color = "white" if self.indice % 2 else "gray90"
			
			self.a = cuentas_apt[self.indice][6]
			self.b = cuentas_apt[self.indice][12]

			if cuentas_apt[self.indice][6] != None:
				self.a = round(cuentas_apt[self.indice][6], 2)
			
			if cuentas_apt[self.indice][12] != None:
				self.b = round(cuentas_apt[self.indice][12],2)
			
			self.tree.insert("",END, tag=('fuente', color),  iid=self.indice, text= cuentas_apt[self.indice][0], 
			values =( cuentas_apt[self.indice][1],(str(cuentas_apt[self.indice][2])).replace('.',','), 
					(str(cuentas_apt[self.indice][3])).replace('.',','),(str(cuentas_apt[self.indice][4])).replace('.',','),
					(str(cuentas_apt[self.indice][5])).replace('.',','),(str(self.a)).replace('.',','),
					(str(cuentas_apt[self.indice][7])).replace('.',','),(str(cuentas_apt[self.indice][8])).replace('.',','),
					cuentas_apt[self.indice][9],(str(cuentas_apt[self.indice][10])).replace('.',','), 
					(str(cuentas_apt[self.indice][11])).replace('.',','), (str(self.b)).replace('.',',')))

			self.indice = self.indice +1

						
	
		except:
			messagebox.showwarning("ADVERTENCIA","Ocurrió un error al mostrar")
			print("m")
			pass

	def sumar_restar_columnas(self):
		print("sumar y restar columnas")
		self.my_connexion = sqlite3.connect("Admicon.db")
		cursor = self.my_connexion.cursor()

		try:
			
			cursor.execute("UPDATE Cuentas_por_apartamento SET  Deuda_actual = ( Deudas_mes_pasado + Recibo + Mora + Alquiler_estacionamiento)")
			cursor.execute("UPDATE Cuentas_por_apartamento SET  Saldo = (Deuda_actual - Pago_Bs - Dolar_a_Bs)")
			self.my_connexion.commit()

			self.datos = cursor.execute("SELECT Deuda_actual, Saldo, id FROM Cuentas_por_apartamento")
			
			for self.dato in self.datos:
				self.deuda_actual = round(self.dato[0], 2)
				self.saldo = round(self.dato[1], 2)
				self.id = self.dato[2]
				print(self.deuda_actual, self.saldo, self.id)
				cursor.execute("UPDATE Cuentas_por_apartamento SET  Deuda_actual='%s' WHERE id='%s'" % (self.deuda_actual, self.id))
				cursor.execute("UPDATE Cuentas_por_apartamento SET  Saldo='%s' WHERE id='%s'" % (self.saldo, self.id))
				print("funciono!")
				self.my_connexion.commit()
				pass

		except:
			messagebox.showwarning("ADVERTENCIA","Ocurrió un error al sumar columnas")
			pass
    
	def añadir_monto_recibo(self):
		self.my_connexion = sqlite3.connect("Admicon.db")
		cursor = self.my_connexion.cursor()
		print(1)
		
		try:
			self.midato = cursor.execute("""SELECT * FROM Total_apto""")
			for row in self.midato:
				self.m_recibo = row[1]

			self.m_recibo_coma = (str(self.m_recibo)).replace('.',',')
			self.total_pago_mostrar = tk.Label(self, text= str(self.m_recibo_coma), font=("Arial"), bg="lightblue", relief=tk.SUNKEN).place(x=120, y=60, width=150)
			
			cursor.execute("""UPDATE Cuentas_por_apartamento SET Recibo='%s'""" % (self.m_recibo))
			print(2)
			self.my_connexion.commit()

			
			self.sumar_restar_columnas()
			print(19)
			self.mostrar()
			print(20)
			self.limpiar_campos_cobro()

		except:
			messagebox.showwarning("ADVERTENCIA","Ocurrió un error al actualizar el registro")
			pass


	
	def poner_recibo_monto(self):
		self.recibo_monto= Recibo_factura.get_monto_recibo()
		print(self.recibo_monto)

		self.my_connexion = sqlite3.connect("Admicon.db")
		cursor = self.my_connexion.cursor()

		
	
				
				
		


class Datos_propietarios(tk.Frame):

	controlbd = Controlador()

	def __init__(self, parent, controller):

		tk.Frame.__init__(self, parent)

		self.config(bg="light blue", width=2440, height=300)

		self.frame_fondo = tk.Frame(self)
		self.frame_fondo.pack(expand=True)
		self.frame_fondo.config(bg="white", width=1440, height=300)

		self.frame_1 = tk.Frame(self.frame_fondo)
		self.frame_1.pack(fill=tk.BOTH, side=tk.TOP)
		self.frame_1.config(bg="light blue", width=1440, height=300)

		self.label_a= tk.Label(self.frame_1, bg="light blue", relief=tk.SUNKEN)
		self.label_a.place(x=10,y=40, width=280, height= 250)

		self.label_b= tk.Label(self.frame_1,bg="light blue", relief=tk.SUNKEN)
		self.label_b.place(x=295,y=70, width=360, height= 220)

		self.label_c= tk.Label(self.frame_1,bg="light blue", relief=tk.SUNKEN)
		self.label_c.place(x=660,y=70, width=360, height= 220)

			#Buttons
		self.button_recibo = tk.Button(self, text="Recibos", command = lambda:controller.show_frame(Recibo_factura), width=20, height=1 ).place(x = 0, y = 0)      
		self.button_cuentas = tk.Button(self, text="Control de Cuentas", command = lambda:controller.show_frame(Control_cuentas),width=20, height=1).place(x = 210, y = 0)     
			
		self.button_correos = tk.Button(self, text="Envío Correos", command = lambda:controller.show_frame(Correos_envio), width=20, height=1).place(x = 630, y = 0) 
		self.button_nomina = tk.Button(self, text="Nómina", 
		command = lambda:controller.show_frame(Nomina_empleados), state= DISABLED, width=20,
								height=1).place(x=840, y=0)

		self.actualizar_b = tk.Button(self, text="Actualizar", command = self.actualizar, width=8, height=1).place(x = 930, y = 250)  

		self.buscar = tk.Button(self, text="Buscar", command = self.busqueda, width=10, height=1).place(x = 810, y = 38)  


			#Labels
		self.label_button_propietarios = tk.Label(self, text="Datos Propietarios", font=("Arial"), bg="lightblue", width=20, height=1 ).place(x = 420, y = 0) 

		self.modificar_datos= tk.Label(self, text="Modificar Datos", font=("Arial"), bg="lightblue" ).place(x=15, y=33)   
		
		self.nombre_apellido = tk.Label(self, text="Nombre y Apellido:", font=("Arial"), bg="lightblue" ).place(x=13, y=70)  
		self.cedula = tk.Label(self, text="Cédula:", font=("Arial"), bg="lightblue" ).place(x=13, y=120)
		self.telefono = tk.Label(self, text="Teléfono:", font=("Arial"), bg="lightblue" ).place(x=13, y=170)
		self.correo = tk.Label(self, text="Correo:", font=("Arial"), bg="lightblue" ).place(x=13, y=220)


		self.vehiculo= tk.Label(self, text="¿Posee vehículo?:", font=("Arial"), bg="lightblue" ).place(x=300, y=80) 
		self.placa = tk.Label(self, text="Placa:", font=("Arial"), bg="lightblue" ).place(x=300, y=130)
		self.habitantes = tk.Label(self, text="Habitantes por apto.:", font=("Arial"), bg="lightblue" ).place(x=300, y=180)
		self.propietario_inquilino = tk.Label(self, text="Propietario o inquilino:", font=("Arial"), bg="lightblue" ).place(x=300, y=230)
		
		self.n_ref= tk.Label(self, text="Nombre del inquilino:", font=("Arial"), bg="lightblue" ).place(x=665, y=80)
		self.n_ref= tk.Label(self, text="Nombre de referencia:", font=("Arial"), bg="lightblue" ).place(x=665, y=130) 
		self.t_ref = tk.Label(self, text="Teléfono de referencia:", font=("Arial"), bg="lightblue" ).place(x=665, y=180)
		self.c_ref = tk.Label(self, text="Correo de referencia.:", font=("Arial"), bg="lightblue" ).place(x=665, y=230)

			#Entry
		
		
		
		self.nombre_apellido_var = StringVar()
		self.cedula_entry_var = StringVar()
		self.telefono_entry_var = StringVar()
		self.correo_var = StringVar()
		self.vehiculo_var = StringVar()
		self.placa_var = StringVar()
		self.habitantes_var = StringVar()
		self.propietario_inquilino_var = StringVar()
		self.nombre_inquilino_var = StringVar()
		self.n_ref_var = StringVar()
		self.t_ref_var = StringVar()
		self.c_ref_var = StringVar()

		self.buscar_entry_var = StringVar()
		



		
		self.nombre_apellido_entry = tk.Entry(self, textvariable= self.nombre_apellido_var).place(x=15, y=95, width=250)
		self.cedula_entry= tk.Entry(self, textvariable= self.cedula_entry_var).place(x=15, y=145, width=250)
		self.telefono_entry = tk.Entry(self, textvariable= self.telefono_entry_var).place(x=15, y=195, width=250)
		self.correo_entry = tk.Entry(self, textvariable= self.correo_var).place(x=15, y=245, width=250)

		self.vehiculo_entry = tk.Entry(self, textvariable= self.vehiculo_var).place(x=300, y=105, width=250)
		self.placa_entry= tk.Entry(self, textvariable= self.placa_var).place(x=300, y=155, width=250)
		self.habitantes_entry = tk.Entry(self, textvariable= self.habitantes_var).place(x=300, y=205, width=250)
		self.propietario_inquilino_entry = tk.Entry(self, textvariable= self.propietario_inquilino_var).place(x=300, y=255, width=250)

		self.nombre_inquilino_entry = tk.Entry(self, textvariable= self.nombre_inquilino_var).place(x=665, y=105, width=250)
		self.n_ref_entry= tk.Entry(self, textvariable= self.n_ref_var).place(x=665, y=155, width=250)
		self.t_ref_entry = tk.Entry(self, textvariable= self.t_ref_var).place(x=665, y=205, width=250)
		self.c_ref_entry = tk.Entry(self, textvariable= self.c_ref_var).place(x=665, y=255, width=250)
		
		self.buscar = tk.Label(self, text="Buscar", font=("Arial"),
							bg="lightblue", width=8, height=1).place(x=300, y=40)

		self.buscar_entry = tk.Entry(self, textvariable= self.buscar_entry_var).place(x=380, y=43, width=420)

		# Frame del treeview
		self.frame_treeview = tk.Frame(self.frame_fondo)
		self.frame_treeview.pack(fill=tk.BOTH, side=tk.BOTTOM)
		self.frame_treeview.config(bg="white", width=1440, height=300)


				# Set the treeview
		self.tree = ttk.Treeview(self.frame_treeview)

		self.treexscroll = tk.Scrollbar(self.frame_treeview, orient=tk.HORIZONTAL)
		self.treexscroll.pack(fill=tk.X, side=tk.BOTTOM)
		# configurar scrollbar
		self.treexscroll.config(command=self.tree.xview)

		self.treeyscroll = tk.Scrollbar(self.frame_treeview, orient=tk.VERTICAL)
		self.treeyscroll.pack(fill=tk.Y, side=tk.RIGHT)
		# configurar scrollbar
		self.treeyscroll.config(command=self.tree.yview)

		# TREEVIEW
		self.tree.config(xscrollcommand=self.treexscroll.set, yscrollcommand=self.treeyscroll.set,
		columns=(
				"col1", "col2", "col3", "col4", "col5", "col6", "col7", "col8", "col9", "col10", "col11", "col12"))

	
		self.tree.column("#0", width=150, stretch= False)
		self.tree.column("col1", width=150, stretch= False)
		self.tree.column("col2", width=150, stretch= False)
		self.tree.column("col3", width=150, stretch= False)
		self.tree.column("col4", width=160, stretch= False)
		self.tree.column("col5", width=150, stretch= False)
		self.tree.column("col6", width=150, stretch= False)
		self.tree.column("col7", width=150, stretch= False)
		self.tree.column("col8", width=150, stretch= False)
		self.tree.column("col9", width=150, stretch= False)
		self.tree.column("col10", width=150, stretch= False)
		self.tree.column("col11", width=150, stretch= False)
		self.tree.column("col12", width=150, stretch= False)
	

		
		self.tree.heading("#0", text="Apartamento", anchor=tk.CENTER)
		self.tree.heading("col1", text="Propietario", anchor=tk.CENTER)
		self.tree.heading("col2", text="Cédula", anchor=tk.CENTER)
		self.tree.heading("col3", text="Teléfono", anchor=tk.CENTER)
		self.tree.heading("col4", text="Correo", anchor=tk.CENTER)
		self.tree.heading("col5", text="Vehículo poseedor", anchor=tk.CENTER)
		self.tree.heading("col6", text="Placa", anchor=tk.CENTER)
		self.tree.heading("col7", text= "Habitantes por apto.", anchor=tk.CENTER)
		self.tree.heading("col8", text="Propietario o inquilino", anchor=tk.CENTER)
		self.tree.heading("col9", text="Nombre de inquilino", anchor=tk.CENTER)
		self.tree.heading("col10", text="Nombre referencial", anchor=tk.CENTER)
		self.tree.heading("col11", text="Tlfno Referencial", anchor=tk.CENTER)
		self.tree.heading("col12", text="Correo referencial", anchor=tk.CENTER)
		

		self.tree.pack()
		self.treeview = self.tree
		
		self.id = 0
		self.iid = 0

		datos_apt = self.controlbd.consulta_datos_apto()
		self.indice = 0
		
		for row in datos_apt:		
			self.tree.tag_configure("gray90", background="gray90")
			self.tree.tag_configure("white", background="white")
			color = "white" if self.indice % 2 else "gray90"

			self.tree.insert("",END, tag=('fuente', color), iid=self.indice, text= row[1], values =(row[2], row[3], 
			row[4],row[5], row[6],row[7], row[8],row[9],
		 	row[10], row[11], row[12], row[13] ))
			self.indice= self.indice+1

	
		
		self.tree.bind('<<TreeviewSelect>>', self.seleccionarUsandoClick)
		self.tree.bind('<<TreeviewSelect>>', self.bindings)
		#self.frame_fondo.bind('<Return>', self.busqueda)
		#self.bindings()
		
	

	def mostrar(self): #Actualizar treeview luego de modificar
		self.my_connexion = sqlite3.connect("Admicon.db")
		cursor = self.my_connexion.cursor()
		registros = self.tree.get_children()
		for elemento in registros:
			self.tree.delete(elemento)
		try:
			datos_apt = cursor.execute("SELECT * FROM Datos_por_apartamento")
			self.indice = 0
			
			for row in datos_apt:			
				self.tree.tag_configure("gray90", background="gray90")
				self.tree.tag_configure("white", background="white")
				color = "white" if self.indice % 2 else "gray90"
				
				self.tree.insert("",END, tag=('fuente', color), iid=self.indice, text = row[1], values =(row[2], 
				row[3],row[4], row[5],row[6], row[7],row[8], row[9],
				row[10], row[11], row[12], row[13] ))
				self.indice= self.indice+1
				
				
	
		except:
			pass


	def limpiarCampos(self):
		self.nombre_apellido_var.set("")
		self.cedula_entry_var.set("")
		self.telefono_entry_var.set("")
		self.correo_var.set("")
		self.vehiculo_var.set("")
		self.placa_var.set("")
		self.habitantes_var.set("")
		self.propietario_inquilino_var.set("")
		self.nombre_inquilino_var.set("")
		self.n_ref_var.set("")
		self.t_ref_var.set("")
		self.c_ref_var.set("")
	
	def seleccionarUsandoClick(self, event):
		item = self.tree.identify('item', event.x, event.y)
		self.nombre_apellido_var.set(self.tree.item(item, "values")[0])
		self.cedula_entry_var.set(self.tree.item(item, "values")[1])
		self.telefono_entry_var.set(self.tree.item(item, "values")[2])
		self.correo_var.set(self.tree.item(item, "values")[3])
		self.vehiculo_var.set(self.tree.item(item, "values")[4])
		self.placa_var.set(self.tree.item(item, "values")[5])
		self.habitantes_var.set(self.tree.item(item, "values")[6])
		self.propietario_inquilino_var.set(self.tree.item(item, "values")[7])
		self.nombre_inquilino_var.set(self.tree.item(item, "values")[8])
		self.n_ref_var.set(self.tree.item(item, "values")[9])
		self.t_ref_var.set(self.tree.item(item, "values")[10])
		self.c_ref_var.set(self.tree.item(item, "values")[11])
		print("you clicked on", self.tree.item(item,"text"))
		self.id_c = self.tree.item(item,"text")
		#self.tree.selection_set('0')
			
	def busqueda(self):
		self.my_connexion = sqlite3.connect("Admicon.db")
		cursor = self.my_connexion.cursor()
		
		try:
			self.criterio = self.buscar_entry_var.get()
			self.criterio = "%s" % self.criterio +"%"
			self.datos = cursor.execute("SELECT id FROM Datos_por_apartamento WHERE Apartamento LIKE '%s' OR Propietario LIKE '%s'" % (self.criterio, self.criterio))
			for row in self.datos:
				self.row_id = row[0]-1
			self.tree.selection_set(self.tree.tag_has(self.row_id)) # move selection
			self.tree.selection_set(self.row_id) # move selection
			self.tree.focus(self.row_id) # move focus
			self.tree.see(self.row_id) # scroll to show it
        	
			
		except:
			messagebox.showwarning("ADVERTENCIA","Ocurrió un error de búsqueda")
			pass


	def bindings(self, event):
		self.tree.bind("<Button-1>", self.seleccionarUsandoClick)


	def actualizar(self):
		self.my_connexion = sqlite3.connect("Admicon.db")
		cursor = self.my_connexion.cursor()
		#datos_apt = cursor.execute("SELECT * FROM Datos_por_apartamento")
		#for row in datos_apt:
		#	print(row)

		try:
			self.s1= self.nombre_apellido_var.get()
			self.s2= self.cedula_entry_var.get()
			self.s3=self.telefono_entry_var.get()
			self.s4=self.correo_var.get()
			self.s5=self.vehiculo_var.get()
			self.s6=self.placa_var.get()
			self.s7=self.habitantes_var.get()
			self.s8=self.propietario_inquilino_var.get()
			self.s9=self.nombre_inquilino_var.get()
			self.s10=self.n_ref_var.get()
			self.s11=self.t_ref_var.get()
			self.s12=self.c_ref_var.get()
			
			cursor.execute("""UPDATE Datos_por_apartamento
			             SET Propietario= ? , CI=?, Tlfno=?, Correo=?, 
						 Vehiculo=?, Placa=?, Habitantes_apto=?, Prop_inq=?, 
						 Inquilino_nombre=?, Nombre_ref=?, Tlfno_ref=?, Correo_ref=? WHERE Apartamento=?""",
						 (self.s1, self.s2, self.s3, self.s4, self.s5, self.s6, self.s7, 
						 self.s8, self.s9, self.s10, self.s11, self.s12, self.id_c))

			self.my_connexion.commit()

		except:
			messagebox.showwarning("ADVERTENCIA","Ocurrió un error al actualizar el registro")
			pass
		self.limpiarCampos()
		self.mostrar()
		
	

class Correos_envio(tk.Frame):

	def __init__(self, parent, controller):

		tk.Frame.__init__(self, parent)
		self.config(bg="light blue", width=2440, height=300)

		
		self.individual_correo = tk.Label(self, text="INDIVIDUAL", font=("Arial"), bg="lightblue" ).place(x=10, y=40) 
		self.apartamento = tk.Label(self, text="Apartamento:", font=("Arial"), bg="lightblue" ).place(x=10, y=100)  

		self.global_correo = tk.Label(self, text="GLOBAL", font=("Arial"), bg="lightblue" ).place(x=500, y=40) 

		self.label_button_correo = tk.Label(self, text="Envío Correos", font=("Arial"), bg="lightblue", width=20, height=1 ).place(x = 630, y = 0) 

			#Entry
		self.apartamento_entry = tk.Entry(self).place(x=150, y=98, width=200) 

			#Buttons
		self.button_recibo = tk.Button(self, text="Recibos", command = lambda:controller.show_frame(Recibo_factura),width=20,
								height=1).place(x=0, y=0)
		self.button_cuentas = tk.Button(self, text="Control de Cuentas", command = lambda:controller.show_frame(Control_cuentas), width=20, height=1).place(x = 210, y = 0)     
		self.button_propietarios = tk.Button(self, text="Datos Propietarios", command = lambda:controller.show_frame(Datos_propietarios),width=20, height=1).place(x = 420, y = 0) 
			
		self.button_nomina = tk.Button(self, text="Nómina", 
		command = lambda:controller.show_frame(Nomina_empleados), state= DISABLED, width=20,
								height=1).place(x=840, y=0)

			#checkbuttons
		self.cea = tk.StringVar() #variable var entero
		self.c = tk.Checkbutton(self, text= "Enviar recibo", variable= self.cea, onvalue = "On", offvalue= "Off", font=("Arial"), bg="lightblue")
		self.c.deselect()
		self.c.place(x=30, y=150)

		self.ceb = tk.StringVar() #variable var entero
		self.c = tk.Checkbutton(self, text= "Añadir Recordatorio de Morosidad", variable= self.ceb, onvalue = "On", offvalue= "Off", font=("Arial"), bg="lightblue")
		self.c.deselect()
		self.c.place(x=30, y=180)

		self.cec = tk.StringVar() #variable var entero
		self.c = tk.Checkbutton(self, text= "Enviar Recibo a Todos los Propietarios", variable= self.cec, onvalue = "On", offvalue= "Off", font=("Arial"), bg="lightblue")
		self.c.deselect()
		self.c.place(x=450, y=180)

		self.ced = tk.StringVar() #variable var entero
		self.c = tk.Checkbutton(self, text= "Enviar Recordatorio de Morosidad a Todos los Morosos", variable= self.ced, onvalue = "On", offvalue= "Off", font=("Arial"), bg="lightblue")
		self.c.deselect()
		self.c.place(x=450, y=280)


class Recibo_factura(tk.Frame):

	controlbd = Controlador()
	
	total_apto = 0,00

	def __init__(self, parent, controller):

		tk.Frame.__init__(self, parent)

		self.config(bg="light blue", width=2440, height=300)
		
		#BUTTONS                         
		self.button_cuentas = tk.Button(self, text="Control de Cuentas", command = lambda:controller.show_frame(Control_cuentas),
		 width=20, height=1).place(x = 210, y = 0)     
		self.button_propietarios = tk.Button(self, text="Datos Propietarios",command = lambda:controller.show_frame(Datos_propietarios), width=20, height=1).place(x = 420, y = 0) 
		self.button_correos = tk.Button(self, text="Envío Correos", command = lambda:controller.show_frame(Correos_envio), width=20, height=1).place(x = 630, y = 0) 
		self.button_nomina = tk.Button(self, text="Nómina", command = lambda:controller.show_frame(Nomina_empleados), 
		state= DISABLED, width=20,height=1).place(x=840, y=0)                        
		self.crear =  tk.Button(self, text="Crear", command= self.crear_factura_general, width=4, height=1).place(x = 410, y = 80)

		self.button_añadir =  tk.Button(self, text="Añadir", command = self.anadir, width=15, height=1).place(x = 40, y = 260)
		self.button_actualizar =  tk.Button(self, text="Actualizar", command = self.actualizar, width=15, height=1).place(x = 200, y = 260)
		self.eliminar =  tk.Button(self, text="Eliminar", command= self.borrar, width=15, height=1).place(x = 360, y = 260)

		self.guardar =  tk.Button(self, text="Guardar", command = self.guardar_recibo, width=25, height=1).place(x = 160, y = 400)

		self.refrescar =  tk.Button(self, text="R", command = self.R, width=20, height=1).place(x = 510, y = 448, width=30)

		self.aceptar_bcv =  tk.Button(self, text="Aceptar", command= self.poner_precio_dolar_BCV, width=20, height=1).place(x = 550, y = 448, width=50)
		self.aceptar_libre =  tk.Button(self, text="Aceptar", command= self.poner_precio_dolar, width=20, height=1).place(x = 550, y = 498, width=50)



		#Labels
		self.label_button_recibo = tk.Label(self, text="Recibos", font=("Arial"), bg="lightblue", width=20, height=1 ).place(x = 0, y = 0)   

		self.vista_recibo= tk.Label(self,text="Vista preliminar recibo:", font=("Arial"), bg="lightblue" ).place(x=10, y=40)
		self.nuevo_recibo= tk.Label(self,text="Nuevo recibo:", font=("Arial"), bg="lightblue" ).place(x=10, y=80)
		self.añadir_item= tk.Label(self,text="Añadir ítems:", font=("Arial"), bg="lightblue" ).place(x=10, y=120)
		self.seccion= tk.Label(self,text="Sección:", font=("Arial"), bg="lightblue" ).place(x=10, y=140)
		self.descripcion= tk.Label(self,text="Descripción:", font=("Arial"), bg="lightblue" ).place(x=10, y=180)
		self.monto= tk.Label(self,text="Monto:", font=("Arial"), bg="lightblue" ).place(x=10, y=220)
		self.total_gen= tk.Label(self, text="Total General:", font=("Arial"), bg="lightblue" ).place(x=10, y=300)
		self.total_apartamento= tk.Label(self, text="Total Apartamento:", font=("Arial"), bg="lightblue" ).place(x=10, y=350)


		self.fecha = tk.Label(self,text="Fecha:", font=("Arial"), bg="lightblue" ).place(x=10, y=450)
		self.fecha_mostrar = tk.Label(self,text= datetime.now().strftime('%d-%m-%Y'), font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=70, y=450, width=120)

		self.dolar_bcv = tk.Label(self,text="Dólar BCV:", font=("Arial"), bg="lightblue" ).place(x=250, y=450)
		
		self.establecer_monto = tk.Label(self,text="Establecer Monto:", font=("Arial"), bg="lightblue" ).place(x=200, y=500)
		
		self.establecer_monto_mostrar_var = StringVar()
		self.establecer_monto_mostrar = tk.Entry(self, textvariable=self.establecer_monto_mostrar_var).place(x=350, y=500, width=150)

		self.monto_establecido = tk.Label(self,text="Monto Establecido:", font=("Arial"), bg="lightblue" ).place(x=700, y=500)
		self.monto_establecido_mostrar = tk.Label(self,text="0,0", font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=850, y=500, width=150)


		#RADIO BUTTONS
		#Tipos de gasto
		self.seleccion = tk.IntVar()
		self.seleccion.set(value=1)
		Radiobutton(self, text="Gastos Ordinarios", variable= self.seleccion, value=1, bg="lightblue" ).place(x=100, y=140)
		Radiobutton(self, text="Anticipo", variable= self.seleccion, value=2, bg="lightblue" ).place(x=250, y=140)
		Radiobutton(self, text="Previsión", variable= self.seleccion, value=3, bg="lightblue" ).place(x=400, y=140)
		Radiobutton(self, text="Gastos Variables", variable= self.seleccion, value=4, bg="lightblue" ).place(x=100, y=160)
		Radiobutton(self, text="Gastos extraordinarios", variable= self.seleccion, value=5, bg="lightblue" ).place(x=250, y=160)
		Radiobutton(self, text="Otros", variable= self.seleccion, value=6, bg="lightblue" ).place(x=400, y=160)
		#Monto(Bs o $USD)
		#m = tk.IntVar()
		#Radiobutton(self, text="Bs", variable=m, value=1, bg="lightblue" ).place(x=450, y=220)
		#Radiobutton(self, text="$", variable=m, value=2, bg="lightblue" ).place(x=400, y=220)

		#Entry
		self.descripcion_var = StringVar()
		self.monto_var = StringVar()
		

		self.vista_recibo = tk.Entry(self).place(x=190, y=45, width=200)
		self.nuevo_recibo = tk.Entry(self).place(x=190, y=80, width=200)
		
		self.descripcion = tk.Entry(self, textvariable=self.descripcion_var).place(x=190, y=180, width=200)
		self.monto = tk.Entry(self, textvariable=self.monto_var).place(x=190, y=220, width=200)


		self.frame_fondo = tk.Frame(self)
		self.frame_fondo.place(x=500, y= 45, height=400, width=525)
		self.frame_fondo.config(bg="green")

		# Frame del treeview
		self.frame_treeview = tk.Frame(self.frame_fondo)
		self.frame_treeview.pack(fill=tk.BOTH, expand="True")
		self.frame_treeview.config(bg="blue")


				# Set the treeview
		self.tree = ttk.Treeview(self.frame_treeview)

		self.treexscroll = tk.Scrollbar(self.frame_treeview, orient=tk.HORIZONTAL)
		self.treexscroll.pack(fill=tk.X, side=tk.BOTTOM)
		# configurar scrollbar
		self.treexscroll.config(command=self.tree.xview)

		self.treeyscroll = tk.Scrollbar(self.frame_treeview, orient=tk.VERTICAL)
		self.treeyscroll.pack(fill=tk.Y, side=tk.RIGHT)
		# configurar scrollbar
		self.treeyscroll.config(command=self.tree.yview)

		# TREEVIEW
		self.tree.config(xscrollcommand=self.treexscroll.set, yscrollcommand=self.treeyscroll.set,
		columns=(
				"col1", "col2"))

	
		self.tree.column("#0", width= 171, stretch= False)
		self.tree.column("col1", width=171, stretch= False)
		self.tree.column("col2", width=171, stretch= False)

		self.tree.heading("#0", text="Descripción", anchor=tk.CENTER)
		self.tree.heading("col1", text="Cuota general", anchor=tk.CENTER)
		self.tree.heading("col2", text="Cuota por apatamento", anchor=tk.CENTER)		


		self.tree.pack(fill= BOTH, expand= "True")
		self.treeview = self.tree
	
		self.id = 0
		self.iid = 0

		
		self.mostrar()
		self.actualizar_total_secciones()


		#Entry2
		self.monto_establecido_mostrar = tk.Label(self, text= self.total_general_coma, font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=200, y=300, width=150)	
		self.monto_establecido_mostrar = tk.Label(self, text= self.total_apto_coma, font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=200, y=350, width=150)	
		self.monto_establecido_dolar_bcv = tk.Label(self,text= "0,0", font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=850, y=500, width=150)
		self.extrae_dolar_bcv()
		self.mostrar_dolar_basedatos()
		
		self.tree.bind('<<TreeviewSelect>>', self.seleccionarUsandoClick)
		self.tree.bind('<<TreeviewSelect>>', self.bindings)



	def mostrar(self): #Actualizar treeview luego de modificar
		datos_apt = self.controlbd.consulta_datos_recibo()
		self.my_connexion = sqlite3.connect("Admicon.db")
		cursor = self.my_connexion.cursor()
		#datos_apt = cursor.execute("SELECT * FROM Recibo ORDER BY Seccion ASC")

		
		#seccion 1 GASTOS ORDINARIOS
		self.seccion_1 = cursor.execute("SELECT * FROM Recibo WHERE Seccion=1") #Gastos ordinarios
		self.indice=0
		self.s1_general=0
		self.s1_apto=0
		for r in self.seccion_1:
			if r[2]>0: 
				self.tree.tag_configure("grey90", background="grey90")
				color="grey90"
				self.tree.insert("",END, tag=('fuente', color), text= "Gastos Ordinarios")
				color = "white"
				self.tree.insert("",END, tag=('fuente', color), text= r[1], values= ((str(round(r[2], 2))).replace('.',','), (str(round(r[3], 2))).replace('.',',') , r[0]))
				self.s1_general=self.s1_general+round(r[2], 2)
				self.s1_apto=self.s1_apto+round(r[3], 2)	
				
				
				for row in self.seccion_1:
					
					color = "white"		
					self.tree.insert("",END, tag=('fuente', color), text= row[1], values= ((str(round(row[2], 2))).replace('.',','), (str(round(row[3], 2))).replace('.',',') , row[0]))
					self.s1_general=self.s1_general+round(row[2], 2)
					self.s1_apto=self.s1_apto+round(row[3], 2)
					self.indice = self.indice+1
				
			
				cursor.execute("UPDATE Recibo_total SET Total_general1='%s', Total_apto1='%s'  WHERE id=1" % (self.s1_general, self.s1_apto))
				print("General y apto:" + str(self.s1_general ) + " - "+ str(self.s1_apto))
				self.my_connexion.commit()

				self.tree.tag_configure("grey60", background="grey60")
				color="grey60"
				self.s1_general_coma = (str(round(self.s1_general, 2))).replace('.',',')
				self.s1_apto_coma = (str(round(self.s1_apto, 2))).replace('.',',')
				self.tree.insert("",END, tag=('fuente', color), text= "Total Gastos Ordinarios", values= (self.s1_general_coma, self.s1_apto_coma))
	
		#seccion 2 ANTICIPO
		self.seccion_2 = cursor.execute("SELECT * FROM Recibo WHERE Seccion=2") #Anticipo

		self.indice=0
		self.s2_general=0
		self.s2_apto=0
		for r in self.seccion_2:
			if r[2]>0: 
				self.tree.tag_configure("grey90", background="grey90")
				color="grey90"
				self.tree.insert("",END, tag=('fuente', color), text= "Anticipo")
				color = "white"
				self.tree.insert("",END, tag=('fuente', color), text= r[1], values= ((str(round(r[2], 2))).replace('.',','), (str(round(r[3], 2))).replace('.',',') , r[0]))
				self.s2_general=self.s2_general+round(r[2], 2)
				self.s2_apto=self.s2_apto+round(r[3], 2)
					
				for row in self.seccion_2:
					
					color = "white"		
					self.tree.insert("",END, tag=('fuente', color), text= row[1], values= ((str(round(row[2], 2))).replace('.',','), (str(round(row[3], 2))).replace('.',',') , row[0]))
					self.s2_general=self.s2_general+round(row[2], 2)
					self.s2_apto=self.s2_apto+round(row[3], 2)
					self.indice = self.indice+1
				
				cursor.execute("UPDATE Recibo_total SET Total_general2='%s', Total_apto2='%s'  WHERE id=1" % (self.s2_general, self.s2_apto))
				self.my_connexion.commit()

				self.s2_general_coma = (str(round(self.s2_general, 2))).replace('.',',')
				self.s2_apto_coma = (str(round(self.s2_apto, 2))).replace('.',',')
				self.tree.tag_configure("grey60", background="grey60")
				color="grey60"
				self.tree.insert("",END, tag=('fuente', color), text= "Total Anticipo", values= (self.s2_general_coma, self.s2_apto_coma))



			#seccion 3 PREVISIÓN
		self.seccion_3 = cursor.execute("SELECT * FROM Recibo WHERE Seccion=3") #Previsión
		self.indice=0
		self.s3_general=0
		self.s3_apto=0

		for r in self.seccion_3:
			if r[2]>0:
				self.tree.tag_configure("grey90", background="grey90")
				color="grey90"
				self.tree.insert("",END, tag=('fuente', color), text= "Previsión")
				color = "white"
				self.tree.insert("",END, tag=('fuente', color), text= r[1], values= ((str(round(r[2], 2))).replace('.',','), (str(round(r[3], 2))).replace('.',',') , r[0]))
				self.s3_general=self.s3_general+round(r[2], 2)
				self.s3_apto=self.s3_apto+round(r[3], 2)
				
				for row in self.seccion_3:
					color = "white"		

					self.tree.insert("",END, tag=('fuente', color), text= row[1], values= ((str(round(row[2], 2))).replace('.',','), (str(round(row[3], 2))).replace('.',',') , row[0]))
					self.s3_general=self.s3_general+round(row[2], 2)
					self.s3_apto=self.s3_apto+round(row[3], 2)
					self.indice = self.indice+1
				
				self.s3_general= round(self.s3_general, 2)
				self.s3_apto= round(self.s3_apto,2)

				cursor.execute("UPDATE Recibo_total SET Total_general3='%s', Total_apto3='%s'  WHERE id=1" % (self.s3_general, self.s3_apto))
				self.my_connexion.commit()
				
				self.s3_general_coma = (str(round(self.s3_general, 2))).replace('.',',')
				self.s3_apto_coma = (str(round(self.s3_apto, 2))).replace('.',',')
				self.tree.tag_configure("grey60", background="grey60")
				color="grey60"
				self.tree.insert("",END, tag=('fuente', color), text= "Total Previsión", values= (self.s3_general_coma, self.s3_apto_coma))

			

			#seccion 4 GASTOS VARIABLES
		self.seccion_4 = cursor.execute("SELECT * FROM Recibo WHERE Seccion=4") #Gastos variables
		self.indice=0
		self.s4_general=0
		self.s4_apto=0
		for r in self.seccion_4:
			if r[2]>0:
				self.tree.tag_configure("grey90", background="grey90")
				color="grey90"
				self.tree.insert("",END, tag=('fuente', color), text= "Gastos Variables")
				color = "white"
				self.tree.insert("",END, tag=('fuente', color), text= r[1], values= ((str(round(r[2], 2))).replace('.',','), (str(round(r[3], 2))).replace('.',',') , r[0]))
				self.s4_general=self.s4_general+round(r[2], 2)
				self.s4_apto=self.s4_apto+round(r[3], 2)
				
	
				for row in self.seccion_4:
					color = "white"		

					self.tree.insert("",END, tag=('fuente', color), text= row[1], values= ((str(round(row[2], 2))).replace('.',','), (str(round(row[3], 2))).replace('.',',') , row[0]))
					self.s4_general=self.s4_general+round(row[2], 2)
					self.s4_apto=self.s4_apto+round(row[3], 2)
					self.indice = self.indice+1
				
				cursor.execute("UPDATE Recibo_total SET Total_general4='%s', Total_apto4='%s'  WHERE id=1" % (self.s4_general, self.s4_apto))
				self.my_connexion.commit()
				
				self.s4_general_coma = (str(round(self.s4_general, 2))).replace('.',',')
				self.s4_apto_coma = (str(round(self.s4_apto, 2))).replace('.',',')
				self.tree.tag_configure("grey60", background="grey60")
				color="grey60"
				self.tree.insert("",END, tag=('fuente', color), text= "Total Gastos Variables", values= (self.s4_general_coma, self.s4_apto_coma))
			



			#seccion 5 GASTOS EXTRAORDINARIOS
		self.seccion_5 = cursor.execute("SELECT * FROM Recibo WHERE Seccion=5") #Gastos extraordinarios
		self.indice=0
		self.s5_general=0
		self.s5_apto=0

		for r in self.seccion_5:
			if r[2]>0: 
				self.tree.tag_configure("grey90", background="grey90")
				color="grey90"
				self.tree.insert("",END, tag=('fuente', color), text= "Gastos Extraordinarios")
				color = "white"
				self.tree.insert("",END, tag=('fuente', color), text= r[1], values= ((str(round(r[2], 2))).replace('.',','), (str(round(r[3], 2))).replace('.',',') , r[0]))
				self.s5_general=self.s5_general+round(r[2], 2)
				self.s5_apto=self.s5_apto+round(r[3], 2)
					
				
				for row in self.seccion_5:
					color = "white"		

					self.tree.insert("",END, tag=('fuente', color), text= row[1], values= ((str(round(row[2], 2))).replace('.',','), (str(round(row[3], 2))).replace('.',',') , row[0]))
					self.s5_general=self.s5_general+round(row[2], 2)
					self.s5_apto=self.s5_apto+round(row[3], 2)
					self.indice = self.indice+1
				

				cursor.execute("UPDATE Recibo_total SET Total_general5='%s', Total_apto5='%s'  WHERE id=1" % (self.s5_general, self.s5_apto))
				self.my_connexion.commit()

				self.s5_general_coma = (str(round(self.s5_general, 2))).replace('.',',')
				self.s5_apto_coma = (str(round(self.s5_apto, 2))).replace('.',',')
				self.tree.tag_configure("grey60", background="grey60")
				color="grey60"
				self.tree.insert("",END, tag=('fuente', color), text= "Total Gastos Extraordinarios", values= (self.s5_general_coma, self.s5_apto_coma))


			#seccion 6 OTROS
		self.seccion_6 = cursor.execute("SELECT * FROM Recibo WHERE Seccion=6") #Otros
		self.indice=0
		self.s6_general=0
		self.s6_apto=0
		for r in self.seccion_6:
			if r[2]>0: 
				self.tree.tag_configure("grey90", background="grey90")
				color="grey90"
				self.tree.insert("",END, tag=('fuente', color), text= "Otros")
				color = "white"
				self.tree.insert("",END, tag=('fuente', color), text= r[1], values= ((str(round(r[2], 2))).replace('.',','), (str(round(r[3], 2))).replace('.',',') , r[0]))
				self.s6_general=self.s6_general+round(r[2], 2)
				self.s6_apto=self.s6_apto+round(r[3], 2)
					
				for row in self.seccion_6:
					color = "white"		

					self.tree.insert("",END, tag=('fuente', color), text= row[1], values= ((str(round(row[2], 2))).replace('.',','), (str(round(row[3], 2))).replace('.',',') , row[0]))
					self.s6_general=self.s6_general+round(row[2], 2)
					self.s6_apto=self.s6_apto+round(r[3], 2)
					self.indice = self.indice+1
				
				self.s6_general_coma = (str(round(self.s6_general, 2))).replace('.',',')
				self.s6_apto_coma = (str(round(self.s6_apto, 2))).replace('.',',')
				self.tree.tag_configure("grey60", background="grey60")
				color="grey60"
				self.tree.insert("",END, tag=('fuente', color), text= "Total Otros", values= (self.s6_general_coma, self.s6_apto_coma))
		
		cursor.execute("UPDATE Recibo_total SET Total_general6='%s', Total_apto6='%s'  WHERE id=1" % (self.s6_general, self.s6_apto))
		self.my_connexion.commit()

		self.total_general= self.s1_general+self.s2_general+self.s3_general+self.s4_general+self.s5_general+self.s6_general
		self.total_apto= self.s1_apto + self.s2_apto + self.s3_apto + self.s4_apto + self.s5_apto + self.s6_apto
		self.total_general= round(self.total_general,2)
		self.total_apto= round(self.total_apto,2)
		
		self.total_general_coma = (str(round(self.total_general, 2))).replace('.',',')
		self.total_apto_coma = (str(round(self.total_apto, 2))).replace('.',',')

		self.tree.tag_configure("salmon1", background="salmon1")
		color="salmon1"
		self.tree.insert("",END, tag=('fuente', color), text= "TOTAL", values= (self.total_general_coma, self.total_apto_coma))
		
		cursor.execute("UPDATE Recibo_total SET Suma_total_general='%s', Suma_total_apto='%s'  WHERE id=1" % (self.total_general, self.total_apto))
		self.my_connexion.commit()

		#Entry
		self.monto_establecido_mostrar = tk.Label(self, text= self.total_general_coma, font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=200, y=300, width=150)	
		self.monto_establecido_mostrar = tk.Label(self, text= self.total_apto_coma, font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=200, y=350, width=150)
		self.dolar_bcv_mostrar = tk.Label(self,text= "0,0", font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=350, y=450, width=150)
		

	def get_monto_recibo(self):
		recibo_total_apto = round(self.total_apto,2)
		self.my_connexion = sqlite3.connect("Admicon.db")
		self.cursor = self.my_connexion.cursor()
		self.cursor.execute("UPDATE Total_apto SET Total='%s' WHERE id='%s'" % (recibo_total_apto,1))
		self.my_connexion.commit()

		
	
	
	def guardar_recibo(self):
		try:
			self.get_monto_recibo()
			#TOTAL POR APARTAMENTO
			self.__class__.total_apto = self.total_apto_coma
			messagebox.showinfo("ADVERTENCIA","Se guardó el recibo con éxito")

		except:
			messagebox.showinfo("ADVERTENCIA","Error al guardar recibo")


	
	
	
	def limpiar_treeview(self): 
		self.tree.delete(*self.tree.get_children())

	
	
	def radio_buttom_seleccionado(self):
		if self.seleccion.get()==1:
			self.rb_variable = 1
		elif self.seleccion.get()==2:
			self.rb_variable = 2
		elif self.seleccion.get()==3:
			self.rb_variable = 3
		elif self.seleccion.get()==4:
			self.rb_variable = 4
		elif self.seleccion.get()==5:
			self.rb_variable = 5
		elif self.seleccion.get()==6:
			self.rb_variable = 6
		


	def anadir(self):
		self.my_connexion = sqlite3.connect("Admicon.db")
		self.cursor = self.my_connexion.cursor()
		print(type(self.monto_var.get()))
		print(self.monto_var.get())
		try:
			print(1)
			#self.monto_individual = (float(self.monto_var.get())*(0.368))%100
			self.monto_glob= float(self.monto_var.get().replace(',','.'))
			self.monto_individual = (self.monto_glob)/272 #272 son la cantidad de apartamentos de capri
			print(2)
			self.monto_individual = round(self.monto_individual, 2)
			print(3)
			self.rb_variable = 0
			print(self.rb_variable)
			self.radio_buttom_seleccionado()
			print(self.seleccion.get())
			print(self.rb_variable)
			self.datos = self.descripcion_var.get(), self.monto_glob, self.monto_individual, self.rb_variable
			print(3333)
			self.cursor.execute("INSERT INTO Recibo VALUES (NULL,?,?,?,?)", (self.datos))
			print(4444)
			self.my_connexion.commit()

			self.limpiarCampos()
			self.limpiar_treeview()
			self.mostrar()
			self.actualizar_total_secciones()
			
			

		except:
			messagebox.showwarning("ADVERTENCIA","Ocurrió un error al crear el registro, verifique conexion con base de datos")
			pass
		
		

	def limpiarCampos(self):
		self.descripcion_var.set("")
		self.monto_var.set("")

	def seleccionarUsandoClick(self, event):
		self.my_connexion = sqlite3.connect("Admicon.db")
		cursor = self.my_connexion.cursor()
		item = self.tree.identify('item', event.x, event.y)
		self.descripcion_var.set(self.tree.item(item, "text"))
		self.monto_var.set(self.tree.item(item, "values")[0])
		
		self.valor = cursor.execute("SELECT Seccion FROM Recibo WHERE id='%s'" % (self.tree.item(item, "values")[2]))
		for ro in self.valor:
			print(ro[0])
		self.seleccion.set(value=ro[0])
		print("you clicked on", self.tree.item(item, "values")[2])
		self.id_c = self.tree.item(item, "values")[2]
		#self.tree.selection_set('0')
	
	def bindings(self, event):
		self.tree.bind("<Button-1>", self.seleccionarUsandoClick)

	def actualizar(self):
		self.my_connexion = sqlite3.connect("Admicon.db")
		cursor = self.my_connexion.cursor()

		try:
			self.monto_glob_n= float(self.monto_var.get().replace(',','.'))
			self.monto_individual_n = (self.monto_glob_n)/272 #272 son la cantidad de apartamentos de capri
			print(2)
			self.monto_individual_n = round(self.monto_individual_n, 2)
			print(3)
			self.rb_variable_n = 0
			self.rb_variable_n = self.seleccion.get()
			print(self.seleccion.get())
			self.radio_buttom_seleccionado()
			print(self.rb_variable)
			self.datos = self.descripcion_var.get(), self.monto_glob_n, self.monto_individual_n, self.rb_variable_n, self.id_c
			print(self.datos)
			print(3333)
			
			cursor.execute("""UPDATE Recibo
			             SET Descripcion= ? , Cuota_general=?, Cuota_apto=?, Seccion=? WHERE id=?""",
						 (self.datos))
						 
			print("d")
			self.my_connexion.commit()
			print("e")
			self.limpiarCampos()
			print("f")
			self.limpiar_treeview()
			print("g")
			self.mostrar()
			print("h")

		except:
			messagebox.showwarning("ADVERTENCIA","Ocurrió un error al actualizar el registro")
			pass
		
	def actualizar_total_secciones(self):
		self.my_connexion = sqlite3.connect("Admicon.db")
		self.cursor = self.my_connexion.cursor()

		for n in range(1,7):
			self.cuota_general = 0
			self.cuota_apto = 0
			self.cuotas = self.cursor.execute("SELECT Cuota_general, Cuota_apto FROM Recibo WHERE Seccion='%s'" % (n))
			for self.cuota in self.cuotas:
					self.cuota_general = self.cuota_general + round(self.cuota[0], 2)
					self.cuota_apto = self.cuota_apto + round(self.cuota[1], 2)
			print("cuota general: " + str(self.cuota_general))
			print("cuota apto: " + str(self.cuota_apto))
			if n ==1:
				self.cursor.execute("UPDATE Recibo_total SET Total_general1='%s', Total_apto1='%s' WHERE id=1" % (round(self.cuota_general, 2), round(self.cuota_apto, 2)))
				self.my_connexion.commit()
				print("1 hecho")
			if n ==2:
				self.cursor.execute("UPDATE Recibo_total SET Total_general2='%s', Total_apto2='%s' WHERE id=1" % (round(self.cuota_general,2), round(self.cuota_apto,2)))
				self.my_connexion.commit()
				print("2 hecho")
			if n ==3:
				self.cursor.execute("UPDATE Recibo_total SET Total_general3='%s', Total_apto3='%s' WHERE id=1" % (round(self.cuota_general,2), round(self.cuota_apto,2)))
				self.my_connexion.commit()
				print("3 hecho")
			if n ==4:
				self.cursor.execute("UPDATE Recibo_total SET Total_general4='%s', Total_apto4='%s' WHERE id=1" % (round(self.cuota_general,2), round(self.cuota_apto,2)))
				self.my_connexion.commit()
				print("4 hecho")
			if n ==5:
				self.cursor.execute("UPDATE Recibo_total SET Total_general5='%s', Total_apto5='%s' WHERE id=1" % (round(self.cuota_general,2), round(self.cuota_apto,2)))
				self.my_connexion.commit()
				print("5 hecho")
			if n ==6:
				self.cursor.execute("UPDATE Recibo_total SET Total_general6='%s', Total_apto6='%s' WHERE id=1" % (round(self.cuota_general,2), round(self.cuota_apto,2)))
				self.my_connexion.commit()
				print("6 hecho")



	def borrar(self):
		self.my_connexion = sqlite3.connect("Admicon.db")
		self.cursor = self.my_connexion.cursor()
		print(self.id_c)
		try:
			if messagebox.askyesno(message="¿Realmente desea eliminar el registro?", title="ADVERTENCIA"):
				self.cursor.execute("DELETE FROM Recibo WHERE id="+self.id_c)
				self.my_connexion.commit()
		except:
			messagebox.showwarning("ADVERTENCIA","Ocurrió un error al tratar de eliminar el registro")
			pass
		self.limpiarCampos()
		self.limpiar_treeview()
		self.mostrar()
		self.actualizar_total_secciones()

	def mostrar_dolar_basedatos(self):
		try:
			self.my_connexion = sqlite3.connect("Admicon.db")
			self.cursor = self.my_connexion.cursor()

			self.precio_dolar = self.cursor.execute("SELECT Dolar_escogido FROM Total_apto WHERE id=1")
			for self.precio_dolar_i in self.precio_dolar:
				self.precio_dolar_i = (str(self.precio_dolar_i[0])).replace('.',',')
			
			self.monto_establecido_mostrar = tk.Label(self,text= self.precio_dolar_i , font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=850, y=500, width=150)
		
		except:
			messagebox.showwarning("ADVERTENCIA","Ocurrió un error")
	
	def poner_precio_dolar(self):
		try:
			self.my_connexion = sqlite3.connect("Admicon.db")
			self.cursor = self.my_connexion.cursor()

			self.precio = (str(self.establecer_monto_mostrar_var.get())).replace(',','.')
			self.precio = float(self.precio)
			self.precio = round(self.precio, 2)
			
			self.cursor.execute("UPDATE Total_apto SET Dolar_manual='%s' WHERE id='%s'" % (self.precio,1))
			self.cursor.execute("UPDATE Total_apto SET Dolar_escogido='%s' WHERE id='%s'" % (self.precio,1))
			self.my_connexion.commit()
			self.monto_establecido_mostrar = tk.Label(self,text= (str(self.precio)).replace('.',','), font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=850, y=500, width=150)
		except:
			#messagebox.showwarning("ADVERTENCIA","Ocurrió un error")
			pass

	
	def poner_precio_dolar_BCV(self):
		try:
				request = requests.get("http://www.google.com", timeout=5)
		except(requests.ConnectionError, requests.Timeout):
				messagebox.showwarning("ADVERTENCIA","No hay conexión a Internet, coloque el monto del dólar manualmente")
		else:
		
			try:
				self.my_connexion = sqlite3.connect("Admicon.db")
				self.cursor = self.my_connexion.cursor()

				
				self.precio_dolar = self.cursor.execute("SELECT Dolar_BCV FROM Total_apto WHERE id=1")
				for self.precio_dolar_i in self.precio_dolar:
					self.precio_dolar_a = self.precio_dolar_i[0]
					self.precio_dolar_i = (str(self.precio_dolar_i[0])).replace('.',',')
				print(self.precio_dolar_a)
				self.cursor.execute("UPDATE Total_apto SET Dolar_escogido='%s' WHERE id='%s'" % (self.precio_dolar_a, 1))
				self.my_connexion.commit()
				print(6)
				self.monto_establecido_mostrar = tk.Label(self,text= self.precio_dolar_i, font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=850, y=500, width=150)
				print(7)
				pass
			
			except:
				messagebox.showwarning("ADVERTENCIA","Ocurrió un error")
				pass

	def R(self):
		try:
			request = requests.get("http://www.google.com", timeout=5)
		except(requests.ConnectionError, requests.Timeout):
				messagebox.showwarning("ADVERTENCIA","No hay conexión a Internet, coloque el monto del dólar manualmente")
		else:
			self.extrae_dolar_bcv()


	def extrae_dolar_bcv(self):
			
			try:
				request = requests.get("http://www.bcv.org.ve", timeout=5)
			except(requests.ConnectionError, requests.Timeout):
				print("No hay conexión a Internet, coloque el monto del dólar manualmente")
			else:

				self.my_connexion = sqlite3.connect("Admicon.db")
				self.cursor = self.my_connexion.cursor()
				URL = "http://www.bcv.org.ve"

				# Realizamos la petición a la web
				req = requests.get(URL)	

				# Comprobamos que la petición nos devuelve un Status Code = 200
				status_code = req.status_code
				print(1)
				if status_code == 200:
					print(1)
					# Pasamos el contenido HTML de la web a un objeto BeautifulSoup()
					print(2)
					soup = BeautifulSoup(req.content, 'html.parser')
					# Obtenemos todos los divs donde están las entradas
					self.entradas = soup.find_all('div', class_= 'col-sm-6 col-xs-6 centrado')
				
				for self.i in self.entradas:
					strong_find = self.i.find('strong')

				colocar_precio = strong_find.text
				colocar_precio = (str(colocar_precio)).strip()
				colocar_precio_float = float(colocar_precio.replace(',','.'))

				#Precio dólar float:
				precio_dolar_float = round(colocar_precio_float, 2)

				#Precio dolar string:
				precio_dolar_str = str(precio_dolar_float).replace('.',',')

				self.cursor.execute("UPDATE Total_apto SET Dolar_BCV='%s', Fecha_dolar_BCV= '%s' WHERE id=1" % (precio_dolar_float,  datetime.now().strftime('%d-%m-%Y')))
				print(5)
				self.my_connexion.commit()
				
				self.dolar_bcv_mostrar = tk.Label(self,text= precio_dolar_str, font=("Arial"), bg="lightblue", relief=tk.SUNKEN ).place(x=350, y=450, width=150)
				pass

	def crear_factura_general_excel(self):
		book = Workbook()
		#book.save("FACTURA GENERAL-'%s'" % (datetime.now().strftime('%m-%Y')))
		s = book.active #Crear hoja
		fecha= datetime.now().strftime('%d-%m-%Y')
		if datetime.now().strftime('%m') == "01":
			print("enero")
			mes = "Enero"
		if datetime.now().strftime('%m') == "02":
			print("Feb")
			mes = "Febrero"
		if datetime.now().strftime('%m') == "03":
			print("MArzo")
			mes = "Marzo"
		if datetime.now().strftime('%m') == "04":
			print("Abril")
			mes = "Abril"
		if datetime.now().strftime('%m') == "05":
			print("Mayo")
			mes = "Mayo"
		if datetime.now().strftime('%m') == "06":
			print("Junio")
			mes = "Junio"
		if datetime.now().strftime('%m') == "07":
			print("Julio")
			mes = "Julio"
		if datetime.now().strftime('%m') == "08":
			print("Ag")
			mes = "Agosto"
		if datetime.now().strftime('%m') == "09":
			print("Spt")
			mes = "Septiembre"
		if datetime.now().strftime('%m') == "10":
			print("Oct")
			mes = "Octubre"
		if datetime.now().strftime('%m') == "11":
			print("Nov")
			mes = "Noviembre"
		if datetime.now().strftime('%m') == "12":
			print("Dic")
			mes = "Diciembre"
		
		fuente0 = Font(name = 'Century', size = 8)

		#TAMAÑOS DE LAS CELDAS
		s.column_dimensions['A'].width = 40
		s.column_dimensions['B'].width = 15
		s.column_dimensions['C'].width = 15
        #QUITAR LINEAS GRISES DE CUADRICULA
		s.sheet_view.showGridLines = False
		
		#ESTABLECER LÍNEAS NEGRAS EN NEGRITA
		bd = Side(style = 'thin', color = "000000")
		bm = Side(style = 'medium', color = "000000")  
       

		print("Head")
		#HEAD RECIBO
		s['A3'] = "            RECIBO DE CONDOMINIO"
		s['A3'].font = fuente0
		
		s['A1'] = "Conjunto residencial Capri"
		s['A1'].font = fuente0

		s['A2'] = "junta de Condominio"
		s['A2'].font = fuente0

		s['A7'] = "Alicuota"
		s['A7'].font = fuente0

		s['B4'] = "Recibo Nro"
		s['B4'].font = fuente0

		s['B5'] = "Mes"
		s['B5'].font = fuente0

		s['B6'] = "Fecha"
		s['B6'].font = fuente0

		s['B7'] = "Alicuota general"
		s['B7'].font = fuente0

		s['C4'] = fecha
		s['C4'].font = fuente0

		s['C5'] = mes
		s['C5'].font = fuente0

		s['C6'] = "inmediato"
		s['C6'].font = fuente0

		s['C7'] = "0,368 %"
		s['C7'].font = fuente0
		
		s['A8'] = "Descripción"
		s['A8'].font = fuente0

		s['B8'] = "Gastos"
		s['B8'].font = fuente0

		s['C8'] = "Cuota por Apartamento"
		s['C8'].font = fuente0
		
		#Lineas Head
		self.lineas_head_top = 'A2', 'B2', 'C2'
		
		for linea in self.lineas_head_top:
			print(linea)
			s['%s' % (linea)].border = Border(bottom = bm)

		self.lineas_head_right = 'A4', 'A5', 'A6', 'A7', 'A8', 'B4', 'B5', 'B6', 'B7', 'B8'
		
		for linea in self.lineas_head_right:
			print(linea)
			s['%s' % (linea)].border = Border(right = bd)

		self.lineas_head_left = 'D3', 'D4', 'D5', 'D6', 'D7', 'D8'
		
		for linea in self.lineas_head_left:
			print(linea)
			s['%s' % (linea)].border = Border(left = bm)
		
		self.lineas_head_bottom = 'A3', 'B3', 'C3', 'A7', 'B7', 'C7', 'A8', 'B8', 'C8' 
		
		for linea in self.lineas_head_bottom:
			print(linea)
			s['%s' % (linea)].border = Border(bottom = bd)

		s['B7'].border = Border(bottom = bd, left = bd)
		s['C7'].border = Border(bottom = bd, left = bd)
       

		#Body
		self.my_connexion = sqlite3.connect("Admicon.db")
		self.cursor = self.my_connexion.cursor()
		
		self.i = 9 #contador celdas
		self.c = 1 #contador
		for self.numero in range(1,7):
			print("número: " + str(self.numero))
			if self.numero ==1:
				seccion="Gastos Ordinarios"
				self.c= 1
			if self.numero ==2:
				seccion="Anticipo"
				self.c =3
			if self.numero ==3:
				seccion="Previsión"
				self.c= 5
			if self.numero ==4:
				seccion="Gastos Variables"
				self.c= 7
			if self.numero ==5:
				seccion="Gastos Extraordinarios"
				self.c = 9
			if self.numero ==6:
				seccion="Otros"
				self.c = 11
			
			#SECCIONES
			
			fuente1 = Font(name = 'Century', italic=True, size = 8, color='FF8000')
			fuente2 = Font(name = 'Century', bold=True, size = 8, color='00FF0000')
			
			self.total_recibo = self.cursor.execute("SELECT * FROM Recibo_total WHERE id=1")
			for segmento in self.total_recibo:
				self.segm = segmento[self.numero*2]
			
			if self.segm > 0:
				self.recibo = self.cursor.execute("SELECT * FROM Recibo WHERE Seccion='%s'" % (self.numero))
				
				self.lineas_head_bottom = 'A', 'B'
				for linea in self.lineas_head_bottom:
					print(linea)
					s['%s' % (linea)+ str(self.i)].border = Border(top= bd, bottom = bd)
				s['C' + str(self.i)].border = Border(top = bd, bottom = bd, right= bm)

				s['A'+ str(self.i)] = seccion
				s['A'+ str(self.i)].font = fuente1

				self.lineas_head_bottom = 'A', 'B'
				for linea in self.lineas_head_bottom:
					print(linea)
					s['%s' % (linea)+ str(self.i)].border = Border(top= bd, bottom = bd)
				s['C' + str(self.i)].border = Border(top= bd, bottom = bd, right= bm)

				self.i= self.i+1

				for self.item in self.recibo:
						print(self.item)
						s['A'+ str(self.i)] = self.item[1]
						s['A'+ str(self.i)].font = fuente0
						s['B'+ str(self.i)] = self.item[2]
						s['B'+ str(self.i)].font = fuente0
						s['B'+ str(self.i)].alignment = Alignment(horizontal='left')
						s['C'+ str(self.i)] = self.item[3]
						s['C'+ str(self.i)].font = fuente0
						s['C'+ str(self.i)].alignment = Alignment(horizontal='left')

						s['C' + str(self.i)].border = Border(right= bm)

						self.i= self.i+1
						

				self.total_recibo = self.cursor.execute("SELECT * FROM Recibo_total WHERE id=1")
				
				for self.t_r in self.total_recibo:
					print(self.t_r)
					s['A'+ str(self.i)] = "Total " + str(seccion)
					s['A'+ str(self.i)].font = fuente1

					s['B'+ str(self.i)] = self.t_r[self.c]
					s['B'+ str(self.i)].font = fuente1
					s['B'+ str(self.i)].alignment = Alignment(horizontal='left')
					self.c= self.c+1
					s['C'+ str(self.i)] = self.t_r[self.c]
					s['C'+ str(self.i)].font = fuente1
					s['C'+ str(self.i)].alignment = Alignment(horizontal='left')
					
					self.lineas_head_bottom = 'A', 'B'
					for linea in self.lineas_head_bottom:
						print(linea)
						s['%s' % (linea)+ str(self.i)].border = Border(top= bd, bottom = bd)
					s['C' + str(self.i)].border = Border(top = bd, bottom = bd, right= bm)


					self.c = self.c+1
					self.i= self.i+1
		
        #FINAL
		self.total_recibo = self.cursor.execute("SELECT * FROM Recibo_total WHERE id=1")
		for self.t_r in self.total_recibo:
			s['A'+ str(self.i)] = "Total General: "	
			s['A'+ str(self.i)].font = fuente2

			s['B'+ str(self.i)] = self.t_r[13]
			s['B'+ str(self.i)].font = fuente2
			s['B'+ str(self.i)].alignment = Alignment(horizontal='center')
			self.c= self.c+1

			s['C'+ str(self.i)] = self.t_r[14]
			s['C'+ str(self.i)].font = fuente2
			s['C'+ str(self.i)].alignment = Alignment(horizontal='center')

			self.lineas_head_bottom = 'A', 'B'
			for linea in self.lineas_head_bottom:
				print(linea)
				s['%s' % (linea)+ str(self.i)].border = Border(top= bd, bottom = bm)
			s['C' + str(self.i)].border = Border(top = bd, bottom = bm, right= bm)


		#wb1 = load_workbook("FACTURA GENERAL-'%s'" % (datetime.now().strftime('%m-%Y')))
		#ws = wb['sheet']
		
		
		nombre_factura_general = "FACTURA GENERAL " + datetime.now().strftime('%m-%Y') 
		nombre_factura_general_excel = nombre_factura_general + ".xlsx"
		book.save(nombre_factura_general_excel)
		
		
	def crear_factura_general(self):
		'''
		P : portrait (vertical)
		L: landscape (horizontal)

		A4: 210x297mm
		'''
		self.pdf = FPDF(orientation = 'P', unit = 'mm', format='A4')
		self.pdf.add_page()

		self.fecha= datetime.now().strftime('%d-%m-%Y')
		if datetime.now().strftime('%m') == "01":
			print("enero")
			self.mes = "Enero"
		if datetime.now().strftime('%m') == "02":
			print("Feb")
			self.mes = "Febrero"
		if datetime.now().strftime('%m') == "03":
			print("MArzo")
			self.mes = "Marzo"
		if datetime.now().strftime('%m') == "04":
			print("Abril")
			self.mes = "Abril"
		if datetime.now().strftime('%m') == "05":
			print("Mayo")
			self.mes = "Mayo"
		if datetime.now().strftime('%m') == "06":
			print("Junio")
			self.mes = "Junio"
		if datetime.now().strftime('%m') == "07":
			print("Julio")
			self.mes = "Julio"
		if datetime.now().strftime('%m') == "08":
			print("Ag")
			self.mes = "Agosto"
		if datetime.now().strftime('%m') == "09":
			print("Spt")
			self.mes = "Septiembre"
		if datetime.now().strftime('%m') == "10":
			print("Oct")
			self.mes = "Octubre"
		if datetime.now().strftime('%m') == "11":
			print("Nov")
			self.mes = "Noviembre"
		if datetime.now().strftime('%m') == "12":
			print("Dic")
			self.mes = "Diciembre"
		
		#HEAD
		self.pdf.set_font('Arial', '', 8)
		self.pdf.multi_cell(w=170, h=5, txt = "Conjunto Residencial Capri \nJunta de Condominio", border = 0, align = 'L', fill =0)
		
		
		self.pdf.multi_cell(w=170, h=8, txt = "RECIBO DE CONDOMINIO", border = 1, align = 'C', fill =0)

		self.pdf.text(x=15, y=65, txt= "Alicuota")
		
		self.pdf.cell(w=90, h=40, txt = "", border = 1, align = 'C', fill =0 )
		

		self.pdf.text(x=105, y=35, txt= "Recibo Nro")
		self.pdf.text(x=105, y=45, txt= "Mes")
		self.pdf.text(x=105, y=55, txt= "Fecha")
		self.pdf.text(x=105, y=65, txt= "Alicuota General")

		self.pdf.text(x=145, y=35, txt= self.fecha)
		self.pdf.text(x=145, y=45, txt= self.mes)
		self.pdf.text(x=145, y=55, txt= "Inmediato")
		self.pdf.text(x=145, y=65, txt= "0,368 %")

		self.pdf.cell(w=40, h=40, txt = "", border = 1, align = 'C', fill =0 )
		self.pdf.multi_cell(w=40, h=40, txt = "", border = 1, align = 'C', fill =0 )

		self.pdf.text(x=45, y= 72, txt= "Descripción")
		self.pdf.text(x=110, y=72, txt= "Gastos")
		self.pdf.text(x=150, y=72, txt= "Cuota P/Apto")
		self.pdf.multi_cell(w=170, h=6, txt = "", border = 1, align = 'C', fill =0)

		#BODY
		#SECCIONES

		self.my_connexion = sqlite3.connect("Admicon.db")
		self.cursor = self.my_connexion.cursor()
		
		self.i = 9 #contador celdas
		self.c = 1 #contador
		altura_y = 78
		for self.numero in range(1,7):
			print("número: " + str(self.numero))
			if self.numero ==1:
				seccion="Gastos Ordinarios"
			if self.numero ==2:
				seccion="Anticipo"
			if self.numero ==3:
				seccion="Previsión"
			if self.numero ==4:
				seccion="Gastos Variables"
			if self.numero ==5:
				seccion="Gastos Extraordinarios"
			if self.numero ==6:
				seccion="Otros"
			
			self.total_recibo = self.cursor.execute("SELECT * FROM Recibo_total WHERE id=1")
			
			for segmento in self.total_recibo:
				self.segm = segmento[self.numero*2]
			
			self.pdf.multi_cell(w=170, h=6, txt = seccion, border = 1, align = 'L', fill =0)
			altura_y = altura_y +6
			
			
			if self.segm > 0:
				self.recibo = self.cursor.execute("SELECT * FROM Recibo WHERE Seccion='%s'" % (self.numero))
				self.a = 0
				for self.item in self.recibo:

					item1= self.item[1]
					conteo = len(self.item[1])
					texto_descripcion = ""
					s = 0
					self.a = self.a +6
					for letra in item1:
						if s < 60:
							texto_descripcion = texto_descripcion + letra
							s=s+1
						if s >= 60:
							if s!= " ":
								texto_descripcion = texto_descripcion + letra
								s=s+1
						if s >= 60:
							if s!= " ":
								texto_descripcion = texto_descripcion + "/n"
								altura_y = altura_y +6
								s = 0
					
					item2= self.item[2]
					item3= self.item[3]
					
					
					self.pdf.text(x=105, y=altura_y, txt= str(item2))
					self.pdf.text(x=150, y=altura_y, txt= str(item3))
					altura_y = altura_y +6
				self.pdf.multi_cell(w=170, h=self.a, txt = texto_descripcion, border = 1, align = 'L', fill =0)

				self.total_recibo = self.cursor.execute("SELECT * FROM Recibo_total WHERE id=1")
				
				for self.t_r in self.total_recibo:
					print(self.t_r)
				
				self.pdf.multi_cell(w=170, h=6, txt = "Total " + seccion, border = 1, align = 'L', fill =0)
				self.pdf.text(x=105, y=altura_y, txt= str(self.t_r[self.c]))
				self.c= self.c+1
				self.pdf.text(x=150, y=altura_y, txt= str(self.t_r[self.c]))
				self.c= self.c+1
				altura_y = altura_y +6

				for self.t_r in self.total_recibo:
					print(self.t_r)
				
				self.pdf.multi_cell(w=170, h=6, txt = "Total " + seccion, border = 1, align = 'L', fill =0)
				self.pdf.text(x=105, y=altura_y, txt= str(self.t_r[13]))
				self.pdf.text(x=150, y=altura_y, txt= str(self.t_r[14]))
				altura_y = altura_y +6
			

		self.nombre_factura_apto = "FACTURA" + datetime.now().strftime('%m-%Y') + ".pdf"
		self.pdf.output(self.nombre_factura_apto)
		print("Funcionó!")
		
		
		




class Nomina_empleados(tk.Frame):

	def __init__(self, parent, controller):

		tk.Frame.__init__(self, parent)
		pass
